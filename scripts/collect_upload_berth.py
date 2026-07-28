# -*- coding: utf-8 -*-
"""
TWL Berth Insight — 선석배정현황 일일 적재 스크립트
=====================================================
매일 06:00 Claude 스케쥴러가 저장한 통합 엑셀을 파싱하여 Supabase에 적재한다.

  입력 : C:\\Temp\\AI_SCM\\터미널 스케쥴 수집\\터미널_선석배정현황_통합_YYYYMMDD.xlsx
  출력 : Supabase bs_vessel_calls (해당 수집일 replace) + bs_collect_log 기록

실행 (06:05 스케쥴 등록 권장):
  set SUPABASE_SERVICE_KEY=<service_role key>   ← 환경변수로만 주입, 코드에 넣지 말 것
  python collect_upload_berth.py                ← 오늘 날짜 파일 처리
  python collect_upload_berth.py 20260727       ← 특정 날짜 파일 처리

SQL 생성 모드 (service key 불필요 — Claude 스케쥴러 + Supabase MCP 경로):
  python collect_upload_berth.py --sql          ← REST 대신 upload_berth.sql 파일 생성
  python collect_upload_berth.py 20260727 --sql
  생성된 C:\\Temp\\AI_SCM\\upload_berth.sql 을 Supabase에서 실행하면 적재 완료
  (동일 수집일 삭제 후 삽입 — 재실행 안전)

필요 패키지: pip install openpyxl requests
"""
import sys, os, re, json, glob, datetime

# Windows 콘솔(cp949)에서 한글/특수문자 출력 깨짐·크래시 방지
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import requests

SUPABASE_URL = 'https://kvmyiualdodcvreoqfin.supabase.co'
SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '')
WATCH_DIR = r'C:\Temp\AI_SCM\터미널 스케쥴 수집'

# ---------- 시트별 컬럼 매핑 (PRD 6.3절) ----------
# 값은 str 또는 [후보1, 후보2 …] — 터미널 포털의 헤더 변형(구/신 포맷) 모두 대응.
# inout: '총물량(IN/OUT)' 결합형 → IN=양하/OUT=적하 로 분리. vessel 없는 시트는 voy로 대체.
MAP = {
 'PNIT': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선명','route':'ROUTE','cct':'반입마감시한','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하','lod':'적하','shift':'Shift','status':'상태'},
 'PNC':  {'berth':'선석','carrier':'운항선사','voy':'모선코드','vessel':'모선명','route':'항로','cct':'반입마감일시','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하수량','lod':'선적수량','shift':'Shift'},
 'HPNT': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선명','route':'ROUTE','cct':'반입마감시한','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하','lod':'적하','shift':'Shift','status':'상태'},
 'HJNC': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선박명','route':'항로','cct':'반입마감일시','eta':'입항일시','etd':'출항일시','ws':'작업시작일시','we':'작업완료일시','dis':'양하','lod':'선적','shift':'S/H'},
 'BNCT': {'berth':'선석','carrier':'선사','voy':['모선항차','모선항차(선사항차)'],'vessel':['선명','선명(ROUTE)'],'route':'ROUTE','cct':'반입마감시한','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하','lod':'적하','shift':'Shift','status':'상태'},
 'DGT':  {'berth':'선석','carrier':'선사코드','voy':['모선항차','모선항차(선사항차)'],'vessel':['모선명','모선명(Route)'],'route':'Route','cct':'반입마감시한','eta':'접안예정일시','etd':'출항예정일시','ws':'작업시작시간','we':'작업완료시간','dis':'양하','lod':'적하','shift':'Shift','status':'상태'},
 'GWCT': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선박명','route':'항로','cct':'반입마감일시','eta':'입항일시','etd':'출항일시','ws':'작업시작일시','we':'작업완료일시','dis':'양하','lod':'선적','shift':'S/H'},
 'E1CT': {'berth':'선석','carrier':'선사','voy':['모선항차','모선항차(입항차/출항차)'],'vessel':['선박명','선박명 Bitt','선박명(Bitt)'],'cct':'반입마감시한(작업완료일시)','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하수량','lod':'적하수량','shift':'Shift'},
 'ICON': {'berth':'선석','carrier':'선사','voy':['모선항차 입항차/출항차','모선항차(입항차/출항차)'],'vessel':['선박명 Bitt(M)','선박명(Bitt)'],'cct':'반입마감일시','eta':'접안(예정)일시','etd':'출항(예정)일시','dis':'양하수량(VAN)','lod':'적하수량(VAN)','shift':'Shift','sub':'터미널'},
 # ---- 2026-07-28 확대: 부산북항·기타 터미널 (FR-02) ----
 'HBCT': {'berth':'선석','carrier':'선사','voy':['선사항차(IN/OUT)','선사항차'],'vessel':'선명','route':'Route','cct':'Closing Time','eta':'접안예정일시','etd':'출항예정일시','ws':'작업예정일시','inout':['총물량(IN/OUT)','총물량 IN/OUT']},
 'BPT':  {'sub':'구분','berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선박명','route':'항로','cct':'반입마감일시','eta':'입항일시','etd':'출항일시','we':'작업완료일시','dis':'양하','lod':'선적','shift':'S/H'},
 'BCT':  {'berth':'선석','carrier':'선사','voy':'모선','vessel':'모선명','route':'ROUTE','cct':'CCT','eta':'접안예정시간(ETB)','etd':'출항예정시간(ETD)','dis':'양하','lod':'적하','status':'상태'},
 'KITL': {'berth':'선석','carrier':'선사','voy':['선사항차(IN/OUT)','선사항차'],'vessel':'모선명','route':'Route','cct':'Closing Time','eta':'접안예정일시','etd':'출항예정일시','inout':['총물량(IN/OUT)','총물량 IN/OUT']},
 'PCTC': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'선박명','route':'항로','cct':'반입마감일시','eta':'입항일시','etd':'출항일시','ws':'작업시작일시','we':'작업완료일시','dis':'양하','lod':'선적','shift':'S/H'},
 'PNCT': {'berth':'선석','voy':'모선항차','vessel':'모선항차','eta':'접안(예정)일시','etd':'출항(예정)일시'},
 'DDCT': {'berth':'선석','carrier':'선사','voy':'모선항차','vessel':'모선명','cct':'CCT','eta':'ETB/ATB','etd':'ETD/ATD','dis':'양하','lod':'적하'},
}


def parse_dt(v):
    """엑셀 값 → 'YYYY-MM-DD HH:MM' 또는 None"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d %H:%M')
    s = str(v).strip().replace('/', '-')
    if not s or s == '-':
        return None
    def _valid(y, mo, d, hm):   # 0000-00-00 등 무효 날짜 제거
        try:
            datetime.datetime.strptime('%s-%s-%s %s' % (y, mo, d, hm), '%Y-%m-%d %H:%M')
            return '%s-%s-%s %s' % (y, mo, d, hm)
        except ValueError:
            return None
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})[ T]?(\d{2}:\d{2})?', s)
    if m:
        return _valid(m.group(1), m.group(2), m.group(3), m.group(4) or '00:00')
    m = re.match(r'(\d{2})-(\d{2})-(\d{2})[ T]?(\d{2}:\d{2})?', s)   # YY/MM/DD (2자리 연도, PNCT 등)
    if m:
        return _valid('20' + m.group(1), m.group(2), m.group(3), m.group(4) or '00:00')
    return None


def parse_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(',', '')))
    except ValueError:
        return 0


def norm_status(raw, eta, etd, we, ref):
    """상태 표준화 (PRD 6.4절). ref = 수집 기준시각 'YYYY-MM-DD 06:00'"""
    if raw:
        s = str(raw).strip().upper()
        if s in ('DEPARTED', 'ARRIVED', 'WORKING', 'PLANNED'):
            return s
    if we and we <= ref:
        return 'DEPARTED'
    if etd and etd <= ref:
        return 'DEPARTED'
    if eta and eta <= ref:
        return 'WORKING'
    return 'PLANNED'


def parse_workbook(path, cdate):
    ref = cdate.strftime('%Y-%m-%d 06:00')
    wb = openpyxl.load_workbook(path, data_only=True)
    out, per_terminal = [], {}
    for sheet, m in MAP.items():
        if sheet not in wb.sheetnames:
            per_terminal[sheet] = 'MISSING'
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        # 헤더 후보(리스트) 지원 — 터미널 포털의 헤더 변형(구/신 포맷) 모두 대응
        def _col(v):
            for c in (v if isinstance(v, list) else [v]):
                if c in header:
                    return header.index(c)
            return None
        idx = {}
        for k, v in m.items():
            j = _col(v)
            if j is not None:
                idx[k] = j
        cnt = 0
        for r in rows[1:]:
            if all(v is None for v in r):
                continue
            g = lambda k: (r[idx[k]] if k in idx else None)
            vessel = g('vessel')
            if not vessel or not str(vessel).strip():
                continue
            # 각 시트 하단 '출처: http…, 수집시각: …' 푸터 행 제거 (전 터미널 방어)
            _vv = str(vessel).strip()
            if _vv.startswith('출처') or 'http' in _vv or '수집시각' in _vv:
                continue
            sub_v = str(g('sub')).strip() if g('sub') else None
            if sheet == 'ICON' and sub_v == 'E1':
                continue  # E1CT 시트와 동일 기항 중복 방지
            vessel_s, route = str(vessel).strip(), g('route')
            mm = re.match(r'^(.*?)\(([^()]*)\)\s*$', vessel_s)   # 선명(ROUTE) 결합형 분리
            if mm and not route:
                vessel_s, route = mm.group(1).strip(), mm.group(2).strip()
            if route and re.match(r'^[+-]?\d+(\.\d+)?m$', str(route).strip()):
                route = None  # Bitt 오프셋 등 비항로 값 제거
            eta, etd, we = parse_dt(g('eta')), parse_dt(g('etd')), parse_dt(g('we'))
            out.append({
                'collected_date': cdate.strftime('%Y-%m-%d'),
                'terminal_cd': sheet,
                'sub_terminal': str(g('sub')).strip() if g('sub') else None,
                'berth': str(g('berth')).strip() if g('berth') is not None else None,
                'carrier': str(g('carrier')).strip() if g('carrier') else None,
                'vessel_name': vessel_s,
                'voyage': (str(g('voy')).replace('(null/null)', '').strip() or None) if g('voy') else None,
                'route': str(route).strip() if route else None,
                'cct': parse_dt(g('cct')), 'eta': eta, 'etd': etd,
                'work_start': parse_dt(g('ws')), 'work_end': we,
                # HBCT 등은 '총물량 IN/OUT' 결합형 → IN=양하/OUT=적하 로 분리
                'discharge_qty': (parse_int(str(g('inout')).split('/')[0]) if g('inout') else parse_int(g('dis'))),
                'load_qty': (parse_int(str(g('inout')).split('/')[1]) if (g('inout') and '/' in str(g('inout'))) else parse_int(g('lod'))),
                'shift_qty': parse_int(g('shift')),
                'status': norm_status(g('status'), eta, etd, we, ref),
                'raw': {header[i]: (str(v) if v is not None else None) for i, v in enumerate(r) if i < len(header)},
            })
            cnt += 1
        per_terminal[sheet] = cnt
    return out, per_terminal


def sb(method, path, body=None):
    r = requests.request(method, SUPABASE_URL + path, json=body, timeout=30, headers={
        'apikey': SERVICE_KEY,
        'Authorization': 'Bearer ' + SERVICE_KEY,
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal',
    })
    if r.status_code >= 300:
        raise RuntimeError(f'{method} {path} -> HTTP {r.status_code}: {r.text[:300]}')


def to_tz(rec):
    """timestamp 필드에 KST(+09) 부여"""
    o = dict(rec)
    for k in ('cct', 'eta', 'etd', 'work_start', 'work_end'):
        if o[k]:
            o[k] = o[k] + '+09'
    return o


SQL_OUT = r'C:\Temp\AI_SCM\sql\upload_berth.sql'


def sql_q(v):
    if v is None or v == '':
        return 'null'
    return "'" + str(v).replace("'", "''") + "'"


def sql_ts(v):
    return 'null' if not v else "'" + v + "+09'"


def write_sql(rows, per_terminal, cdate, fname):
    """--sql 모드: REST 대신 실행용 SQL 파일 생성 (동일 수집일 replace)"""
    lines = [
        '-- 자동 생성: collect_upload_berth.py --sql (%s, %d건)' % (cdate, len(rows)),
        "delete from public.bs_vessel_calls where collected_date = '%s';" % cdate,
        'insert into public.bs_vessel_calls',
        '  (collected_date, terminal_cd, sub_terminal, berth, carrier, vessel_name, voyage, route,',
        '   cct, eta, etd, work_start, work_end, discharge_qty, load_qty, shift_qty, status) values',
    ]
    vals = []
    for r in rows:
        vals.append('  (%s)' % ', '.join([
            sql_q(r['collected_date']), sql_q(r['terminal_cd']), sql_q(r['sub_terminal']),
            sql_q(r['berth']), sql_q(r['carrier']), sql_q(r['vessel_name']),
            sql_q(r['voyage']), sql_q(r['route']),
            sql_ts(r['cct']), sql_ts(r['eta']), sql_ts(r['etd']),
            sql_ts(r['work_start']), sql_ts(r['work_end']),
            str(r['discharge_qty']), str(r['load_qty']), str(r['shift_qty']), sql_q(r['status']),
        ]))
    lines.append(',\n'.join(vals) + ';')
    lines.append(
        "insert into public.bs_collect_log (collected_date, file_name, total_rows, per_terminal, status, message) values "
        "('%s', %s, %d, %s::jsonb, 'SUCCESS', 'Claude 스케쥴러 --sql 적재');"
        % (cdate, sql_q(fname), len(rows), sql_q(json.dumps(per_terminal, ensure_ascii=False)))
    )
    with open(SQL_OUT, 'w', encoding='utf-8') as fp:
        fp.write('\n'.join(lines))


def main():
    args = [a for a in sys.argv[1:] if a != '--sql']
    sql_mode = '--sql' in sys.argv[1:]
    if not sql_mode and not SERVICE_KEY:
        sys.exit('환경변수 SUPABASE_SERVICE_KEY 가 설정되지 않았습니다. (키 없이 쓰려면 --sql 모드 사용)')
    ymd = args[0] if args else datetime.date.today().strftime('%Y%m%d')
    cdate = datetime.datetime.strptime(ymd, '%Y%m%d').date()
    pattern = os.path.join(WATCH_DIR, f'터미널_선석배정현황_통합_{ymd}.xlsx')
    files = glob.glob(pattern)
    fname = os.path.basename(files[0]) if files else None

    try:
        if not files:
            raise FileNotFoundError(f'수집 파일 없음: {pattern}')
        rows, per_terminal = parse_workbook(files[0], cdate)
        if not rows:
            raise ValueError('정규화 결과 0건 — 시트 구조 변경 여부 확인 필요')

        if sql_mode:
            write_sql(rows, per_terminal, str(cdate), fname)
            print(f'[OK] {cdate} {len(rows)}건 → {SQL_OUT} 생성 완료 — {json.dumps(per_terminal, ensure_ascii=False)}')
            return

        # 동일 수집일 replace: 삭제 후 삽입 (FR-01)
        sb('DELETE', f'/rest/v1/bs_vessel_calls?collected_date=eq.{cdate}')
        for i in range(0, len(rows), 100):
            sb('POST', '/rest/v1/bs_vessel_calls', [to_tz(x) for x in rows[i:i + 100]])

        sb('POST', '/rest/v1/bs_collect_log', {
            'collected_date': str(cdate), 'file_name': fname, 'total_rows': len(rows),
            'per_terminal': per_terminal, 'status': 'SUCCESS', 'message': None,
        })
        print(f'[OK] {cdate} {len(rows)}건 적재 완료 — {json.dumps(per_terminal, ensure_ascii=False)}')
    except Exception as e:
        try:
            sb('POST', '/rest/v1/bs_collect_log', {
                'collected_date': str(cdate), 'file_name': fname, 'total_rows': 0,
                'per_terminal': None, 'status': 'FAIL', 'message': str(e)[:500],
            })
        except Exception:
            pass  # 로그 기록마저 실패해도 원래 오류를 우선 표시
        sys.exit(f'[FAIL] {e}')


if __name__ == '__main__':
    main()
