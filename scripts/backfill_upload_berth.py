# -*- coding: utf-8 -*-
"""
TWL Berth Insight — 선석배정현황 과거분 백필 스크립트
=====================================================
일일 수집기(collect_upload_berth.py)와 달리 여러 날짜의 통합 엑셀을 한 번에 적재하며,
기간 중 바뀐 터미널 공표 양식(헤더 변형)을 자동 대응한다.

  입력 : 터미널_선석배정현황_통합_YYYYMMDD.xlsx (파일 경로들 또는 글롭 패턴)
  출력 : Supabase bs_vessel_calls (수집일별 replace) + bs_collect_log 기록

사용법:
  python backfill_upload_berth.py --dry-run "C:\\Users\\user\\Downloads\\터미널_선석배정현황_통합_*.xlsx"
      → 파싱 결과만 요약 출력 (업로드 없음, 헤더 미해결 필드 경고 포함)

  python backfill_upload_berth.py --sql "C:\\Users\\user\\Downloads\\터미널_선석배정현황_통합_*.xlsx"
      → sql\\backfill_berth.sql 생성 (service key 불필요 — Supabase MCP/SQL Editor 실행용)

  set SUPABASE_SERVICE_KEY=<service_role key>
  python backfill_upload_berth.py "C:\\Users\\user\\Downloads\\터미널_선석배정현황_통합_*.xlsx"
      → REST로 직접 적재 (수집일별 삭제 후 삽입 — 재실행 안전)

헤더 변형 대응(2026-07-14~28 실측 전수 스캔 기준):
  - 줄바꿈/띄어쓰기 변형: '반입 마감일시', '작업 시작일시', '모선항차\\n입항차/출항차' 등
  - 컬럼 분리↔결합: BNCT·DGT '모선항차'+'선사항차' ↔ '모선항차(선사항차)',
    '선명'+'ROUTE' ↔ '선명(ROUTE)', E1CT·ICON '선박명'+'Bitt' ↔ '선박명(Bitt)'
  - 작업량 통합 컬럼: '양하/적하/Shift' 한 칸에 '618/675/0' 형태 → 3필드 분해
  - 괄호 감싼 일시값: E1CT '(2026-07-14 01:00)' → 괄호 제거 후 파싱
"""
import sys, os, re, glob, json, datetime

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from collect_upload_berth import (MAP, parse_dt, parse_int, norm_status, log_verdict,
                                  to_tz, sb, sql_q, sql_ts)
import openpyxl
from xls_compat import ensure_xlsx

# ---------- 헤더 변형 대응표 (필드별 대체 후보, 공백 정규화 후 비교) ----------
VARIANTS = {
    'HJNC': {'cct': ['반입 마감일시'], 'ws': ['작업 시작일시'], 'we': ['작업 완료일시']},
    'GWCT': {'cct': ['반입 마감일시'], 'eta': ['입항 일시'], 'etd': ['출항 일시'],
             'ws': ['작업 시작일시'], 'we': ['작업 완료일시']},
    'BNCT': {'voy': ['모선항차', '모선항차(선사항차) Head (Bridge) Stern'],
             'vessel': ['선명', '선명 (ROUTE)'],
             'route': ['ROUTE']},
    'DGT':  {'voy': ['모선항차'], 'vessel': ['모선명'], 'carrier': ['선사'],
             'route': ['Route']},
    'E1CT': {'voy': ['모선항차 입항차/출항차', '모선항차'],
             'vessel': ['선박명 Bitt', '선박명 Bitt(M)', '선박명'],
             'cct': ['반입마감시한 (작업완료일시)', '반입마감시한']},
    'ICON': {'voy': ['모선항차 입항차/출항차', '모선항차'],
             'vessel': ['선박명 Bitt(M)', '선박명']},
    'BCT':  {'voy': ['모선/항차']},
    'DDCT': {'voy': ['모선/항차']},
}
# '양하/적하/Shift'가 한 컬럼으로 합쳐진 변형 (값: '618/675/0')
COMBINED_QTY = ['양하/적하/Shift', '작업량 양하/적하/Shift', '작업량 양하 / 적하 / Shift']

# 통합 파일명 → 수집일. 수집기가 .xls(SpreadsheetML)로 저장하는 경우도 있어 두 확장자 모두 허용
FNAME_RE = re.compile(r'터미널_선석배정현황_통합_(\d{8})\.xlsx?$', re.I)


def norm_h(h):
    """헤더 정규화: 공백·줄바꿈 전부 제거 — 공백 위치만 다른 변형('접안예정 일시' 등)을 흡수"""
    return re.sub(r'\s+', '', str(h)) if h is not None else ''


def parse_dt_x(v):
    """괄호로 감싼 일시값('(2026-07-14 01:00)') 대응 후 기존 파서 위임"""
    if isinstance(v, str):
        v = v.strip()
        if v.startswith('(') and v.endswith(')'):
            v = v[1:-1].strip()
    return parse_dt(v)


def build_idx(header, sheet):
    """정규화 헤더 위치 매핑: 현행 MAP(후보 리스트 지원) 우선, 실패 시 VARIANTS 후보 순차 적용"""
    hn = [norm_h(h) for h in header]
    m, v = MAP[sheet], VARIANTS.get(sheet, {})
    idx = {}
    for k, primary in m.items():
        prim = list(primary) if isinstance(primary, list) else [primary]
        for cand in prim + v.get(k, []):
            cn = norm_h(cand)
            if cn in hn:
                idx[k] = hn.index(cn)
                break
    for k, cands in v.items():          # MAP에 없는 보조 필드 (예: BNCT의 분리형 route)
        if k in idx:
            continue
        for cand in cands:
            cn = norm_h(cand)
            if cn in hn:
                idx[k] = hn.index(cn)
                break
    combined = next((hn.index(norm_h(c)) for c in COMBINED_QTY if norm_h(c) in hn), None)
    unresolved = [k for k in m if k not in idx]
    if combined is not None:
        unresolved = [k for k in unresolved if k not in ('dis', 'lod', 'shift')]
    return idx, combined, unresolved


def parse_workbook_v(path, cdate):
    """변형 대응 파서 — 반환 스키마는 일일 수집기와 동일"""
    ref = cdate.strftime('%Y-%m-%d 06:00')
    wb = openpyxl.load_workbook(ensure_xlsx(path), data_only=True)
    out, per_terminal, warns = [], {}, []
    for sheet in MAP:
        if sheet not in wb.sheetnames:
            per_terminal[sheet] = 'MISSING'
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            per_terminal[sheet] = 0
            continue
        header = list(rows[0])
        idx, combined, unresolved = build_idx(header, sheet)
        if unresolved:
            warns.append('%s %s: 미해결 필드 %s' % (cdate, sheet, unresolved))
        cnt = 0
        for r in rows[1:]:
            if all(x is None for x in r):
                continue
            def g(k, _r=r):
                x = _r[idx[k]] if k in idx and idx[k] < len(_r) else None
                if isinstance(x, str):   # 수집 원문의 '@@' 변경표시·잉여 공백 제거
                    x = re.sub(r'\s+', ' ', x.replace('@@', ' ')).strip()
                    return x or None
                return x
            vessel = g('vessel')
            if not vessel or not str(vessel).strip():
                continue
            # 각 시트 하단 '출처: http…, 수집시각: …' 푸터 행 제거 (전 터미널 방어)
            _vv = str(vessel).strip()
            if _vv.startswith('출처') or 'http' in _vv or '수집시각' in _vv:
                continue
            sub_v = str(g('sub')).strip() if g('sub') else None
            if sheet == 'ICON' and sub_v == 'E1':
                continue                      # E1CT 시트와 동일 기항 중복 방지
            vessel_s, route = str(vessel).strip(), g('route')
            mm = re.match(r'^(.*?)\(([^()]*)\)\s*$', vessel_s)   # 선명(ROUTE) 결합형 분리
            if mm and not route:
                vessel_s, route = mm.group(1).strip(), mm.group(2).strip()
            if route and re.match(r'^[+-]?\d+(\.\d+)?m$', str(route).strip()):
                route = None                  # Bitt 오프셋 등 비항로 값 제거
            eta, etd, we = parse_dt_x(g('eta')), parse_dt_x(g('etd')), parse_dt_x(g('we'))
            if combined is not None and combined < len(r) and r[combined] is not None:
                parts = [p.strip() for p in str(r[combined]).split('/')]
                dis = parse_int(parts[0]) if len(parts) > 0 else 0
                lod = parse_int(parts[1]) if len(parts) > 1 else 0
                shf = parse_int(parts[2]) if len(parts) > 2 else 0
            else:
                dis, lod, shf = parse_int(g('dis')), parse_int(g('lod')), parse_int(g('shift'))
            out.append({
                'collected_date': cdate.strftime('%Y-%m-%d'),
                'terminal_cd': sheet,
                'sub_terminal': sub_v,
                'berth': str(g('berth')).strip() if g('berth') is not None else None,
                'carrier': str(g('carrier')).strip() if g('carrier') else None,
                'vessel_name': vessel_s,
                'voyage': (str(g('voy')).replace('(null/null)', '').strip() or None) if g('voy') else None,
                'route': str(route).strip() if route else None,
                'cct': parse_dt_x(g('cct')), 'eta': eta, 'etd': etd,
                'work_start': parse_dt_x(g('ws')), 'work_end': we,
                'discharge_qty': dis, 'load_qty': lod, 'shift_qty': shf,
                'status': norm_status(g('status'), eta, etd, we, ref),
            })
            cnt += 1
        per_terminal[sheet] = cnt
    return out, per_terminal, warns


def collect_files(args):
    """인자(파일/글롭 혼용) → [(date, path)] 날짜 오름차순"""
    paths = []
    for a in args:
        paths.extend(glob.glob(a) if any(c in a for c in '*?') else [a])
    dated = []
    for p in paths:
        m = FNAME_RE.search(os.path.basename(p))
        if not m:
            sys.exit('[FAIL] 파일명에서 날짜를 찾을 수 없습니다: %s' % p)
        if not os.path.exists(p):
            sys.exit('[FAIL] 파일 없음: %s' % p)
        dated.append((datetime.datetime.strptime(m.group(1), '%Y%m%d').date(), p))
    dated.sort()
    return dated


def sql_lines_for(rows, per_terminal, cdate, fname):
    """수집일 1건분 SQL (동일 수집일 replace + 로그)"""
    lines = [
        "delete from public.bs_vessel_calls where collected_date = '%s';" % cdate,
        'insert into public.bs_vessel_calls',
        '  (collected_date, terminal_cd, sub_terminal, berth, carrier, vessel_name, voyage, route,',
        '   cct, eta, etd, work_start, work_end, discharge_qty, load_qty, shift_qty, status) values',
    ]
    vals = []
    for r in rows:
        vals.append('  (%s)' % ', '.join([
            sql_q(r['collected_date']), sql_q(r['terminal_cd']), sql_q(r['sub_terminal']),
            sql_q(r['berth']), sql_q(r['carrier']), sql_q(r['vessel_name']),
            sql_q(r['voyage']), sql_q(r['route']),
            sql_ts(r['cct']), sql_ts(r['eta']), sql_ts(r['etd']),
            sql_ts(r['work_start']), sql_ts(r['work_end']),
            str(r['discharge_qty']), str(r['load_qty']), str(r['shift_qty']), sql_q(r['status']),
        ]))
    lines.append(',\n'.join(vals) + ';')
    lines.append(
        "delete from public.bs_collect_log where collected_date = '%s' and message like '백필%%';" % cdate)
    lines.append(
        "insert into public.bs_collect_log (collected_date, file_name, total_rows, per_terminal, status, message) values "
        "('%s', %s, %d, %s::jsonb, 'SUCCESS', '백필 적재 (backfill_upload_berth.py)');"
        % (cdate, sql_q(fname), len(rows), sql_q(json.dumps(per_terminal, ensure_ascii=False)))
    )
    return lines


SQL_OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'sql', 'backfill_berth.sql')


def main():
    flags = [a for a in sys.argv[1:] if a.startswith('--')]
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    dry, sql_mode = '--dry-run' in flags, '--sql' in flags
    if not args:
        sys.exit('사용법: backfill_upload_berth.py [--dry-run|--sql] <파일|글롭 패턴>...')
    if not dry and not sql_mode and not os.environ.get('SUPABASE_SERVICE_KEY', ''):
        sys.exit('환경변수 SUPABASE_SERVICE_KEY 미설정 — --sql 모드로 SQL 파일을 생성해 MCP/SQL Editor로 적재하십시오.')

    dated = collect_files(args)
    print('대상 파일 %d개: %s ~ %s' % (len(dated), dated[0][0], dated[-1][0]))
    all_sql, total, all_warns = [], 0, []

    for cdate, path in dated:
        fname = os.path.basename(path)
        rows, per_terminal, warns = parse_workbook_v(path, cdate)
        all_warns.extend(warns)
        if not rows:
            print('[SKIP] %s: 정규화 결과 0건 — 확인 필요' % cdate)
            continue
        zero = [k for k, v in per_terminal.items() if v == 0]
        print('%s: %4d건%s' % (cdate, len(rows), ('  (0건 시트: %s)' % ','.join(zero)) if zero else ''))
        total += len(rows)

        if dry:
            continue
        if sql_mode:
            all_sql.append('-- ===== %s (%s, %d건) =====' % (cdate, fname, len(rows)))
            all_sql.extend(sql_lines_for(rows, per_terminal, str(cdate), fname))
            continue
        # REST 직접 적재 (수집일별 replace)
        sb('DELETE', '/rest/v1/bs_vessel_calls?collected_date=eq.%s' % cdate)
        for i in range(0, len(rows), 100):
            sb('POST', '/rest/v1/bs_vessel_calls', [to_tz(x) for x in rows[i:i + 100]])
        st, msg = log_verdict(per_terminal, '백필 적재 (backfill_upload_berth.py)')
        sb('POST', '/rest/v1/bs_collect_log', {
            'collected_date': str(cdate), 'file_name': fname, 'total_rows': len(rows),
            'per_terminal': per_terminal, 'status': st, 'message': msg,
        })

    for w in all_warns:
        print('[WARN] ' + w)
    if sql_mode and all_sql:
        with open(SQL_OUT, 'w', encoding='utf-8') as fp:
            fp.write('\n'.join(['-- 자동 생성: backfill_upload_berth.py --sql (%d개 수집일, 총 %d건)'
                                % (len(dated), total)] + all_sql))
        print('[OK] 총 %d건 → %s 생성 완료' % (total, SQL_OUT))
    elif dry:
        print('[DRY-RUN] 총 %d건 — 업로드 없음' % total)
    else:
        print('[OK] 총 %d건 적재 완료' % total)


if __name__ == '__main__':
    main()
