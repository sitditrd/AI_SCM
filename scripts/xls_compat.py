# -*- coding: utf-8 -*-
"""
엑셀 형식 호환 계층 — SpreadsheetML 2003(.xls) → .xlsx 변환
=====================================================
06시 수집기가 저장하는 통합 파일의 형식이 바뀔 수 있다(2026-08-04 실측:
`.xlsx`(zip) → `.xls`(SpreadsheetML 2003 XML)). openpyxl 은 후자를 읽지 못해
적재가 통째로 건너뛰어졌다. 이 모듈이 형식을 판별해 필요할 때만 변환한다.

  from xls_compat import ensure_xlsx
  path = ensure_xlsx(원본경로)      # xlsx 면 그대로, SpreadsheetML 이면 임시 xlsx 경로

지원 형식
  - PK\\x03\\x04  : 진짜 xlsx (그대로 반환)
  - <?xml … mso-application progid="Excel.Sheet"?> : SpreadsheetML 2003 → 변환
  - \\xd0\\xcf\\x11\\xe0 : 구형 OLE2 xls → 미지원(명시적 오류)
"""
import os
import tempfile
import xml.etree.ElementTree as ET

import openpyxl

SS = 'urn:schemas-microsoft-com:office:spreadsheet'
Q = '{%s}' % SS


def sniff(path):
    """파일 앞부분으로 형식 판별 → 'xlsx' | 'spreadsheetml' | 'ole2' | 'unknown'"""
    with open(path, 'rb') as fp:
        head = fp.read(512)
    if head[:2] == b'PK':
        return 'xlsx'
    if head[:4] == b'\xd0\xcf\x11\xe0':
        return 'ole2'
    if b'mso-application' in head or SS.encode() in head:
        return 'spreadsheetml'
    return 'unknown'


def _cell_value(data_el):
    """<Data ss:Type="…"> → 파이썬 값. 숫자는 int/float 로, 나머지는 문자열."""
    if data_el is None:
        return None
    txt = ''.join(data_el.itertext())
    if txt == '':
        return None
    if data_el.get(Q + 'Type') == 'Number':
        try:
            f = float(txt)
            return int(f) if f.is_integer() else f
        except ValueError:
            return txt
    return txt


def convert_spreadsheetml(src, dst=None):
    """SpreadsheetML 2003 → xlsx 로 변환하고 결과 경로 반환.
    ss:Index(빈 칸 건너뛰기) 속성을 존중해 컬럼이 밀리지 않게 한다."""
    root = ET.parse(src).getroot()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for ws_el in root.findall(Q + 'Worksheet'):
        name = (ws_el.get(Q + 'Name') or 'Sheet')[:31]
        ws = wb.create_sheet(title=name)
        table = ws_el.find(Q + 'Table')
        if table is None:
            continue
        r = 0
        for row_el in table.findall(Q + 'Row'):
            ri = row_el.get(Q + 'Index')
            r = (int(ri) - 1) if ri else r
            r += 1
            c = 0
            for cell_el in row_el.findall(Q + 'Cell'):
                ci = cell_el.get(Q + 'Index')
                c = (int(ci) - 1) if ci else c
                c += 1
                v = _cell_value(cell_el.find(Q + 'Data'))
                if v is not None:
                    ws.cell(row=r, column=c, value=v)
    if dst is None:
        base = os.path.splitext(os.path.basename(src))[0]
        dst = os.path.join(tempfile.gettempdir(), base + '.converted.xlsx')
    wb.save(dst)
    return dst


def ensure_xlsx(path):
    """openpyxl 이 읽을 수 있는 경로를 돌려준다. 필요할 때만 변환한다."""
    kind = sniff(path)
    if kind == 'xlsx':
        return path
    if kind == 'spreadsheetml':
        out = convert_spreadsheetml(path)
        print('[INFO] SpreadsheetML(.xls) 감지 → xlsx 변환: %s' % os.path.basename(out))
        return out
    if kind == 'ole2':
        raise ValueError('구형 OLE2 .xls 는 지원하지 않습니다(수집기에서 xlsx 로 저장 필요): %s' % path)
    raise ValueError('알 수 없는 엑셀 형식: %s' % path)
