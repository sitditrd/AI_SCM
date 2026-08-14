#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
국내 컨테이너 터미널 선석배정현황 자동 수집기
================================================

기능
----
1. 국내 16개 컨테이너 터미널의 선석배정현황 수집
2. 터미널별 Excel 생성
3. 첨부 기준 파일과 동일한 양식의 통합 Excel 생성
4. Outlook 발송용 HTML 본문과 pending JSON 생성
5. 한 터미널이 실패해도 나머지 터미널 계속 처리

실행 환경
---------
- Python 3.10 이상 권장
- Windows 작업 스케줄러에서 실행 가능
- 정적 HTML 사이트: requests + BeautifulSoup
- 동적/프레임/가상 스크롤 사이트: Playwright

실행 환경 준비
--------------
- 일반 실행 시 누락된 Python 패키지와 Playwright Chromium을 자동 설치합니다.
- 사전 설치만 수행하려면 다음 명령을 사용합니다.
    py terminal_schedule_collector.py --setup
- 자동 설치를 금지하려면 --no-auto-install 옵션을 사용합니다.

기본 실행
---------
    py terminal_schedule_collector.py

브라우저 화면을 보면서 점검
----------------------------
    py terminal_schedule_collector.py --headed

일부 터미널만 실행
------------------
    py terminal_schedule_collector.py --terminal PNIT,PNC,HPNT

requests 방식만 실행하고 브라우저 대상은 실패 처리
-----------------------------------------------------
    py terminal_schedule_collector.py --no-browser

출력 경로 변경
--------------
    py terminal_schedule_collector.py --output-root "D:\\터미널 스케쥴 정보"

주의
----
- 실제 Outlook 메일은 발송하지 않습니다.
- _tools\pending\report_YYYYMMDD.json 생성까지만 수행합니다.
- 사이트 DOM이 변경되면 해당 터미널만 실패로 기록되고 전체 작업은 계속됩니다.
- 웹사이트 이용 정책 및 사내 보안 정책을 확인한 뒤 운영하십시오.
"""

from __future__ import annotations

import argparse
import asyncio
import html as html_lib
import importlib
import importlib.util
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
import traceback
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path, PureWindowsPath
from typing import Any, Iterable, Sequence
from urllib.parse import urljoin
try:
    from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
except ImportError:  # Python 3.8 이하 방어용(본 프로그램은 3.10 이상 권장)
    ZoneInfo = None  # type: ignore[assignment]
    ZoneInfoNotFoundError = Exception  # type: ignore[assignment,misc]


def build_kst_timezone():
    """Windows에 IANA 시간대 DB가 없어도 KST를 반환한다.

    Windows Python에는 ``Asia/Seoul`` 데이터가 기본 포함되지 않는 경우가
    있다. 이때 프로그램 시작 전에 ``ZoneInfoNotFoundError``가 발생하므로,
    현재 한국 표준시와 동일한 UTC+09:00 고정 오프셋으로 대체한다.
    ``tzdata`` 패키지가 설치되어 있으면 정식 IANA 시간대를 사용한다.
    """
    if ZoneInfo is not None:
        try:
            return ZoneInfo("Asia/Seoul")
        except (ZoneInfoNotFoundError, ModuleNotFoundError):
            pass
    return timezone(timedelta(hours=9), name="KST")


KST = build_kst_timezone()
APP_VERSION = "2026.08.13.3"
DEFAULT_WINDOWS_ROOT = PureWindowsPath(r"D:\터미널 스케쥴 정보")
ALLOWED_RECIPIENT = "itt@twsc.co.kr"

# 첨부된 통합 엑셀 샘플(2026-08-12)의 출력 양식을 그대로 재현한다.
EXCEL_HEADER_FILL = "D9D9D9"
EXCEL_BORDER_COLOR = "999999"
EXCEL_SOURCE_FONT_COLOR = "666666"
EXCEL_BODY_FONT_SIZE = 10
EXCEL_SOURCE_FONT_SIZE = 9
EXCEL_MIN_COLUMN_WIDTH = 6

HEADER_KEYWORDS = (
    "선석",
    "선사",
    "모선",
    "선명",
    "항차",
    "접안",
    "입항",
    "출항",
    "반입",
    "closing",
    "양하",
    "적하",
    "선적",
    "shift",
    "route",
    "상태",
    "berth",
    "vessel",
    "departure",
    "arrival",
)

FOOTER_WORDS = (
    "다음 페이지",
    "next page",
    "이전 페이지",
    "previous page",
    "페이지 이동",
    "로그인",
    "회원가입",
    "전체메뉴",
)

DATE_PATTERN = re.compile(
    r"(?:20\d{2}[./-]\d{1,2}[./-]\d{1,2})|(?:\d{1,2}[./-]\d{1,2}\s+\d{1,2}:\d{2})"
)
WHITESPACE_PATTERN = re.compile(r"[\t\r\n\u00a0 ]+")
CONTROL_PATTERN = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


class CollectorError(RuntimeError):
    """수집 과정의 예상 가능한 오류."""


# ---------------------------------------------------------------------------
# 실행 환경 자동 점검 및 설치
# ---------------------------------------------------------------------------

# Python 파일을 직접 실행해도 필요한 패키지와 Playwright Chromium을
# 자동으로 준비한다. 작업 스케줄러와 수동 CMD 실행이 같은 Python
# 인터프리터를 사용하도록 모든 설치 명령은 sys.executable 기준으로 수행한다.
RUNTIME_DEPENDENCIES: tuple[tuple[str, str], ...] = (
    ("requests", "requests>=2.31,<3"),
    ("bs4", "beautifulsoup4>=4.12,<5"),
    ("openpyxl", "openpyxl>=3.1,<4"),
    ("tzdata", "tzdata>=2025.2"),
)
PLAYWRIGHT_DEPENDENCY: tuple[str, str] = (
    "playwright",
    "playwright>=1.45,<2",
)


def module_available(module_name: str) -> bool:
    """현재 Python 인터프리터에서 모듈을 import할 수 있는지 확인한다."""
    try:
        return importlib.util.find_spec(module_name) is not None
    except (ImportError, ModuleNotFoundError, ValueError):
        return False


def run_setup_process(
    command: Sequence[str],
    description: str,
    *,
    timeout_seconds: int = 1200,
    capture_output: bool = False,
) -> subprocess.CompletedProcess[str]:
    """설치/점검 하위 프로세스를 실행하고 실패 원인을 보존한다."""
    print(f"[환경점검] {description}", flush=True)
    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    try:
        completed = subprocess.run(
            list(command),
            check=False,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=capture_output,
            timeout=timeout_seconds,
            env=env,
        )
    except subprocess.TimeoutExpired as exc:
        raise CollectorError(
            f"{description} 시간이 {timeout_seconds}초를 초과했습니다."
        ) from exc
    except OSError as exc:
        raise CollectorError(
            f"{description} 실행에 실패했습니다: {type(exc).__name__}: {exc}"
        ) from exc
    return completed


def ensure_python_dependencies(*, include_playwright: bool, auto_install: bool) -> None:
    """누락된 Python 패키지를 현재 인터프리터에 자동 설치한다."""
    dependencies = list(RUNTIME_DEPENDENCIES)
    if include_playwright:
        dependencies.append(PLAYWRIGHT_DEPENDENCY)

    missing = [(module, spec) for module, spec in dependencies if not module_available(module)]
    if not missing:
        print("[환경점검] Python 패키지 확인 완료", flush=True)
        return

    missing_modules = ", ".join(module for module, _ in missing)
    install_specs = [spec for _, spec in missing]
    install_command = [
        sys.executable,
        "-m",
        "pip",
        "install",
        "--disable-pip-version-check",
        "--no-input",
        *install_specs,
    ]

    if not auto_install:
        raise CollectorError(
            "필수 Python 패키지가 없습니다: "
            f"{missing_modules}. 다음 명령을 실행하십시오: "
            + " ".join(install_command)
        )

    completed = run_setup_process(
        install_command,
        f"누락 패키지 자동 설치: {missing_modules}",
        timeout_seconds=1200,
    )
    if completed.returncode != 0:
        raise CollectorError(
            "Python 패키지 자동 설치에 실패했습니다. "
            f"종료 코드={completed.returncode}. 명령: {' '.join(install_command)}"
        )

    importlib.invalidate_caches()
    still_missing = [module for module, _ in missing if not module_available(module)]
    if still_missing:
        raise CollectorError(
            "설치 명령은 완료됐지만 현재 Python에서 모듈을 찾지 못했습니다: "
            + ", ".join(still_missing)
            + f". 사용 중인 Python: {sys.executable}"
        )
    print("[환경점검] Python 패키지 자동 설치 완료", flush=True)


def playwright_chromium_smoke_test() -> tuple[bool, str]:
    """별도 프로세스에서 Chromium을 실제로 실행해 설치 상태를 검증한다."""
    smoke_code = (
        "from playwright.sync_api import sync_playwright\n"
        "with sync_playwright() as p:\n"
        "    browser = p.chromium.launch(headless=True)\n"
        "    browser.close()\n"
    )
    completed = run_setup_process(
        [sys.executable, "-c", smoke_code],
        "Playwright Chromium 실행 확인",
        timeout_seconds=90,
        capture_output=True,
    )
    output = "\n".join(
        value.strip()
        for value in (completed.stdout or "", completed.stderr or "")
        if value and value.strip()
    )
    return completed.returncode == 0, output


def ensure_playwright_chromium(*, auto_install: bool) -> None:
    """Chromium 실행 파일이 없으면 Playwright 공식 설치 명령으로 자동 설치한다."""
    ok, first_output = playwright_chromium_smoke_test()
    if ok:
        print("[환경점검] Playwright Chromium 확인 완료", flush=True)
        return

    install_command = [sys.executable, "-m", "playwright", "install", "chromium"]
    if not auto_install:
        detail = first_output[-1200:] if first_output else "Chromium 실행 실패"
        raise CollectorError(
            "Playwright Chromium이 설치되지 않았거나 실행할 수 없습니다. "
            f"다음 명령을 실행하십시오: {' '.join(install_command)}. 원인: {detail}"
        )

    completed = run_setup_process(
        install_command,
        "Playwright Chromium 자동 설치",
        timeout_seconds=1800,
    )
    if completed.returncode != 0:
        raise CollectorError(
            "Playwright Chromium 자동 설치에 실패했습니다. "
            f"종료 코드={completed.returncode}. 명령: {' '.join(install_command)}"
        )

    ok, second_output = playwright_chromium_smoke_test()
    if not ok:
        detail = second_output[-1600:] if second_output else first_output[-1600:]
        raise CollectorError(
            "Playwright Chromium 설치 후에도 브라우저 실행 점검에 실패했습니다. "
            f"사용 중인 Python: {sys.executable}. 원인: {detail}"
        )
    print("[환경점검] Playwright Chromium 자동 설치 및 실행 확인 완료", flush=True)


def ensure_runtime_environment(
    *,
    include_browser: bool,
    auto_install: bool,
) -> None:
    """실제 수집 전에 전역 의존성을 한 번 검증한다.

    이 점검을 통과하지 못하면 수집을 시작하지 않으므로, Playwright가 없는
    상태에서 requests 대상 4개만 저장하고 불완전한 발송 JSON을 만드는 일을
    방지한다.
    """
    print(f"[환경점검] 수집기 버전: {APP_VERSION}", flush=True)
    print(f"[환경점검] Python: {sys.executable}", flush=True)
    ensure_python_dependencies(
        include_playwright=include_browser,
        auto_install=auto_install,
    )
    if include_browser:
        ensure_playwright_chromium(auto_install=auto_install)
    print("[환경점검] 실행 환경 준비 완료", flush=True)



def quarantine_existing_pending(output_root: str, run_date_value: str | None) -> Path | None:
    """같은 날짜의 기존 발송 지시서를 대피시켜 불완전 파일 발송을 막는다.

    재실행 시 이전 시도에서 생성된 pending JSON이 남아 있으면 작업 스케줄러가
    새 수집 완료 전에 과거 통합 파일을 발송할 수 있다. 삭제하지 않고
    ``_tools/pending_stale``로 이동해 추적 가능하게 보존한다.
    """
    run_date = parse_run_date(run_date_value)
    root = Path(output_root).expanduser().resolve()
    pending = root / "_tools" / "pending" / f"report_{run_date:%Y%m%d}.json"
    if not pending.exists():
        return None

    stale_dir = root / "_tools" / "pending_stale"
    stale_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    target = stale_dir / f"report_{run_date:%Y%m%d}_stale_{timestamp}.json"
    counter = 1
    while target.exists():
        target = stale_dir / (
            f"report_{run_date:%Y%m%d}_stale_{timestamp}_{counter}.json"
        )
        counter += 1
    pending.replace(target)
    print(
        f"[안전조치] 기존 발송 지시서를 대피했습니다: {target}",
        flush=True,
    )
    return target


@dataclass(frozen=True)
class TerminalConfig:
    code: str
    port: str
    name: str
    url: str
    folder: str
    strategy: str
    expected_columns: int | None = None
    request_first: bool = False
    note: str = ""
    menu_texts: tuple[str, ...] = ()


@dataclass
class CellSpec:
    text: str
    is_header: bool = False
    rowspan: int = 1
    colspan: int = 1


@dataclass
class RawTable:
    cell_rows: list[list[CellSpec]]
    source: str = ""
    caption: str = ""


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[str]]
    score: float
    source: str = ""
    diagnostics: dict[str, Any] = field(default_factory=dict)


@dataclass
class TerminalResult:
    config: TerminalConfig
    success: bool
    headers: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)
    error: str = ""
    note: str = ""
    source_url: str = ""
    collected_at: datetime = field(default_factory=lambda: datetime.now(KST))
    output_file: Path | None = None
    method: str = ""

    @property
    def count(self) -> int:
        return len(self.rows)


TERMINALS: tuple[TerminalConfig, ...] = (
    TerminalConfig(
        "PNIT",
        "부산신항",
        "부산신항국제터미널",
        "https://www.pnitl.com/infoservice/vessel/vslScheduleList.jsp",
        "부산신항_부산신항국제터미널(PNIT)",
        "generic_table",
        request_first=True,
    ),
    TerminalConfig(
        "PNC",
        "부산신항",
        "부산신항만",
        "https://svc.pncport.com/info/CMS/Ship/Info.pnc?mCode=MN014",
        "부산신항_부산신항만(PNC)",
        "generic_table",
        request_first=True,
    ),
    TerminalConfig(
        "HPNT",
        "부산신항",
        "HMM PSA 신항만",
        "https://www.hpnt.co.kr/infoservice/vessel/vslScheduleList.jsp",
        "부산신항_HMMPSA신항만(HPNT)",
        "generic_table",
        request_first=True,
    ),
    TerminalConfig(
        "HJNC",
        "부산신항",
        "한진부산컨테이너터미널",
        "https://www.hjnc.co.kr/esvc/vessel/berthScheduleT",
        "부산신항_한진부산컨테이너터미널(HJNC)",
        "generic_table",
    ),
    TerminalConfig(
        "BNCT",
        "부산신항",
        "부산신항컨테이너터미널",
        "https://info.bnctkorea.com/esvc/vessel/berthScheduleT",
        "부산신항_부산신항컨테이너터미널(BNCT)",
        "generic_table",
    ),
    TerminalConfig(
        "DGT",
        "부산신항",
        "동원글로벌터미널부산",
        "https://info.dgtbusan.com/DGT/esvc/vessel/berthScheduleT",
        "부산신항_동원글로벌터미널부산(DGT)",
        "generic_table",
    ),
    TerminalConfig(
        "BCT",
        "부산신항",
        "부산컨테이너터미널",
        "https://info.bct2-4.com/",
        "부산신항_부산컨테이너터미널(BCT)",
        "nexacro",
        expected_columns=18,
        menu_texts=("선석 배정현황(목록)", "선석배정현황(목록)", "선석 배정현황"),
    ),
    TerminalConfig(
        "HBCT",
        "부산북항",
        "한국허치슨터미널",
        "https://custom.hktl.com/jsp/T01/sunsuk.jsp",
        "부산북항_한국허치슨터미널(HBCT)",
        "legacy_hutchison",
    ),
    TerminalConfig(
        "BPT",
        "부산북항",
        "부산항터미널",
        "https://info.bptc.co.kr/",
        "부산북항_부산항터미널(BPT)",
        "bpt",
        menu_texts=("선석배정 현황(T)", "선석배정현황(T)", "선석배정 현황"),
    ),
    TerminalConfig(
        "GWCT",
        "광양항",
        "광양항서부컨테이너터미널",
        "http://www.gwct.co.kr/sub/sub_B2",
        "광양항_광양항서부컨테이너터미널(GWCT)",
        "generic_table",
    ),
    TerminalConfig(
        "KITL",
        "광양항",
        "한국국제터미널",
        "https://info.kitl.com/jsp/T01/sunsuk.jsp",
        "광양항_한국국제터미널(KITL)",
        "legacy_hutchison",
    ),
    TerminalConfig(
        "E1CT",
        "인천항",
        "E1컨테이너터미널",
        "http://www.e1ct.co.kr/info/terminal/berthText",
        "인천항_E1컨테이너터미널(E1CT)",
        "generic_table",
    ),
    TerminalConfig(
        "ICON",
        "인천항",
        "iCON 통합",
        "https://scon.icpa.or.kr/vescall/list.do?menuKey=19",
        "인천항_iCON통합(ICON)",
        "generic_table",
        request_first=True,
        note="사이트 정책에 따라 1페이지 20건만 수집",
    ),
    TerminalConfig(
        "PCTC",
        "평택당진항",
        "평택컨테이너터미날",
        "http://www.pctc21.com/esvc/vessel/berthScheduleT",
        "평택당진항_평택컨테이너터미날(PCTC)",
        "generic_table",
        expected_columns=17,
    ),
    TerminalConfig(
        "PNCT",
        "평택당진항",
        "평택동방아이포트",
        "http://www.pnct.co.kr/infoservice/index.html",
        "평택당진항_평택동방아이포트(PNCT)",
        "nexacro",
        expected_columns=9,
        menu_texts=("선석 스케쥴(텍스트)", "선석 스케줄(텍스트)", "선석스케쥴(텍스트)"),
    ),
    TerminalConfig(
        "DDCT",
        "대산항",
        "동방대산컨테이너터미널",
        "https://ds.dongbang.co.kr/infoservice/index.html",
        "대산항_동방대산컨테이너터미널(DDCT)",
        "nexacro",
        expected_columns=13,
        menu_texts=("선석 배정 현황(텍스트)", "선석배정현황(텍스트)", "선석 배정 현황"),
    ),
)

SHEET_ORDER = [
    "PNIT",
    "PNC",
    "HPNT",
    "HJNC",
    "BNCT",
    "DGT",
    "GWCT",
    "E1CT",
    "ICON",
    "HBCT",
    "BPT",
    "BCT",
    "KITL",
    "PCTC",
    "PNCT",
    "DDCT",
]


# ---------------------------------------------------------------------------
# 공통 유틸리티
# ---------------------------------------------------------------------------


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u200b", "").strip()
    text = WHITESPACE_PATTERN.sub(" ", text)
    return CONTROL_PATTERN.sub("", text).strip()


def safe_excel_text(value: Any) -> str:
    text = normalize_text(value)
    # 외부 데이터의 수식 실행을 막되, 사이트에서 빈 값 대용으로 쓰는 단독 하이픈(-)은
    # 첨부 샘플과 동일하게 그대로 표시한다.
    if text != "-" and text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def display_width(text: str) -> int:
    width = 0
    for char in text:
        width += 2 if ord(char) > 255 else 1
    return width


def unique_headers(headers: Sequence[str]) -> list[str]:
    used: dict[str, int] = {}
    output: list[str] = []
    for index, raw in enumerate(headers, start=1):
        base = normalize_text(raw) or f"컬럼{index}"
        count = used.get(base, 0) + 1
        used[base] = count
        output.append(base if count == 1 else f"{base} ({count})")
    return output


def deduplicate_rows(rows: Iterable[Sequence[str]]) -> list[list[str]]:
    seen: set[tuple[str, ...]] = set()
    output: list[list[str]] = []
    for row in rows:
        normalized = tuple(normalize_text(v) for v in row)
        if not any(normalized):
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        output.append(list(normalized))
    return output


def parse_run_date(value: str | None) -> date:
    if not value:
        return datetime.now(KST).date()
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError("--date는 YYYY-MM-DD 형식이어야 합니다.") from exc


def default_output_root() -> Path:
    if os.name == "nt":
        return Path(str(DEFAULT_WINDOWS_ROOT))
    return Path.cwd() / "터미널 스케쥴 정보"


def to_windows_path(windows_root: PureWindowsPath, relative_path: Path) -> str:
    parts = [part for part in relative_path.parts if part not in (".", "")]
    return str(windows_root.joinpath(*parts))


def sanitize_error(exc: BaseException, limit: int = 700) -> str:
    text = normalize_text(f"{type(exc).__name__}: {exc}")
    return text[:limit]


def configure_logging(log_file: Path, verbose: bool) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    level = logging.DEBUG if verbose else logging.INFO
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    root = logging.getLogger()
    root.setLevel(level)
    root.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(formatter)
    root.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    root.addHandler(console_handler)


# ---------------------------------------------------------------------------
# HTML 표 파싱
# ---------------------------------------------------------------------------


def expand_cell_rows(cell_rows: Sequence[Sequence[CellSpec]]) -> tuple[list[list[str]], list[list[bool]]]:
    """rowspan/colspan을 반영하여 직사각형 행렬로 확장한다."""
    values_grid: list[list[str]] = []
    flags_grid: list[list[bool]] = []
    pending: dict[int, tuple[int, str, bool]] = {}

    for specs in cell_rows:
        row_values: list[str] = []
        row_flags: list[bool] = []
        col = 0

        def consume_pending_at(position: int) -> bool:
            item = pending.get(position)
            if item is None:
                return False
            remaining, text, is_header = item
            while len(row_values) <= position:
                row_values.append("")
                row_flags.append(False)
            row_values[position] = text
            row_flags[position] = is_header
            if remaining <= 1:
                del pending[position]
            else:
                pending[position] = (remaining - 1, text, is_header)
            return True

        for spec in specs:
            while consume_pending_at(col):
                col += 1

            rowspan = max(1, int(spec.rowspan or 1))
            colspan = max(1, int(spec.colspan or 1))
            text = normalize_text(spec.text)

            for offset in range(colspan):
                position = col + offset
                while len(row_values) <= position:
                    row_values.append("")
                    row_flags.append(False)

                # 헤더 colspan은 각 열에 동일한 상위 헤더를 반복한다.
                # 데이터 colspan은 첫 열에만 값을 두고 나머지는 빈 값으로 보존한다.
                cell_text = text if spec.is_header or offset == 0 else ""
                row_values[position] = cell_text
                row_flags[position] = bool(spec.is_header)

                if rowspan > 1:
                    pending[position] = (rowspan - 1, cell_text, bool(spec.is_header))
            col += colspan

        if pending:
            max_pending = max(pending)
            while col <= max_pending:
                if consume_pending_at(col):
                    col += 1
                else:
                    while len(row_values) <= col:
                        row_values.append("")
                        row_flags.append(False)
                    col += 1

        values_grid.append(row_values)
        flags_grid.append(row_flags)

    width = max((len(row) for row in values_grid), default=0)
    for row, flags in zip(values_grid, flags_grid):
        row.extend([""] * (width - len(row)))
        flags.extend([False] * (width - len(flags)))
    return values_grid, flags_grid


def row_header_score(row: Sequence[str], flags: Sequence[bool] | None = None) -> float:
    joined = " ".join(normalize_text(v).lower() for v in row if normalize_text(v))
    keyword_hits = sum(1 for keyword in HEADER_KEYWORDS if keyword in joined)
    th_count = sum(bool(v) for v in (flags or []))
    nonempty = sum(1 for value in row if normalize_text(value))
    date_hits = len(DATE_PATTERN.findall(joined))
    return keyword_hits * 8 + th_count * 2 + min(nonempty, 20) * 0.15 - date_hits * 6


def flatten_headers(
    values: Sequence[Sequence[str]], start: int, end: int, width: int
) -> list[str]:
    headers: list[str] = []
    for col in range(width):
        pieces: list[str] = []
        for row_index in range(start, end + 1):
            value = normalize_text(values[row_index][col])
            if not value:
                continue
            if pieces and value == pieces[-1]:
                continue
            # 상위 헤더가 하위 헤더 문자열에 그대로 포함되면 중복 결합하지 않는다.
            if pieces and (value in pieces[-1] or pieces[-1] in value):
                if len(value) > len(pieces[-1]):
                    pieces[-1] = value
                continue
            pieces.append(value)
        headers.append(" / ".join(pieces))
    return unique_headers(headers)


def is_footer_or_noise(row: Sequence[str]) -> bool:
    nonempty = [normalize_text(v) for v in row if normalize_text(v)]
    if not nonempty:
        return True
    joined = " ".join(nonempty).lower()
    if any(word in joined for word in FOOTER_WORDS):
        return True
    if len(nonempty) <= 2 and re.search(r"(?:^|\s)\[?\d+\]?\s*\[?\d+\]?", joined):
        return True
    if len(nonempty) <= 2 and ("total" in joined or "전체 조회 건수" in joined):
        return True
    return False


def analyze_raw_table(raw: RawTable) -> ParsedTable | None:
    values, flags = expand_cell_rows(raw.cell_rows)
    if len(values) < 2:
        return None
    width = max((len(row) for row in values), default=0)
    if width < 4:
        return None

    search_limit = min(len(values), 12)
    scores = [row_header_score(values[i], flags[i]) for i in range(search_limit)]
    header_end = max(range(search_limit), key=lambda i: scores[i])

    # th 행이 존재하면 키워드 점수와 함께 우선 고려한다.
    th_candidates = [
        i for i in range(search_limit) if sum(flags[i]) >= max(1, width // 5)
    ]
    if th_candidates:
        best_th = max(th_candidates, key=lambda i: scores[i])
        if scores[best_th] >= scores[header_end] - 2:
            header_end = best_th

    header_start = header_end
    # 다단 헤더: 현재 헤더 바로 위의 행이 th 중심이고 날짜가 없으면 포함한다.
    while header_start > 0:
        previous = header_start - 1
        previous_joined = " ".join(values[previous])
        previous_th = sum(flags[previous])
        previous_score = row_header_score(values[previous], flags[previous])
        if DATE_PATTERN.search(previous_joined):
            break
        if previous_th > 0 and previous_score > 1:
            header_start = previous
            continue
        break

    headers = flatten_headers(values, header_start, header_end, width)
    header_keyword_hits = sum(
        1
        for keyword in HEADER_KEYWORDS
        if keyword in " ".join(headers).lower()
    )

    rows: list[list[str]] = []
    date_rows = 0
    for raw_row in values[header_end + 1 :]:
        row = [normalize_text(v) for v in raw_row[:width]]
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        if is_footer_or_noise(row):
            continue
        joined = " ".join(row)
        nonempty_count = sum(bool(v) for v in row)
        if nonempty_count < max(2, width // 5):
            continue
        # 헤더 반복 행 제외
        if sum(1 for value in row if value in headers) >= max(3, width // 2):
            continue
        if DATE_PATTERN.search(joined):
            date_rows += 1
        rows.append(row)

    rows = deduplicate_rows(rows)
    if not rows:
        return None

    numeric_cells = sum(
        1
        for row in rows[:50]
        for value in row
        if re.fullmatch(r"[\d,./:-]+", value or "")
    )
    score = (
        header_keyword_hits * 30
        + min(len(rows), 200) * 1.2
        + min(date_rows, 100) * 3
        + min(width, 30)
        + min(numeric_cells, 100) * 0.1
    )
    if header_keyword_hits < 2:
        score -= 80
    if date_rows == 0:
        score -= 30

    return ParsedTable(
        headers=headers,
        rows=rows,
        score=score,
        source=raw.source,
        diagnostics={
            "width": width,
            "header_start": header_start,
            "header_end": header_end,
            "header_keyword_hits": header_keyword_hits,
            "date_rows": date_rows,
            "raw_rows": len(values),
        },
    )


def choose_best_table(raw_tables: Sequence[RawTable]) -> ParsedTable | None:
    parsed = [table for raw in raw_tables if (table := analyze_raw_table(raw))]
    if not parsed:
        return None
    parsed.sort(key=lambda item: item.score, reverse=True)
    return parsed[0]


def raw_tables_from_html(content: bytes, encoding_hint: str | None, source: str) -> list[RawTable]:
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:  # pragma: no cover - 환경 오류
        raise CollectorError(
            "beautifulsoup4가 설치되지 않았습니다. "
            "py -m pip install beautifulsoup4 를 실행하십시오."
        ) from exc

    encodings: list[str] = []
    if encoding_hint:
        encodings.append(encoding_hint)
    encodings.extend(["utf-8", "cp949", "euc-kr"])

    text: str | None = None
    for encoding in OrderedDict.fromkeys(encodings):
        try:
            candidate = content.decode(encoding, errors="strict")
            # 한글 페이지인데 대체문자가 많으면 다른 인코딩을 시도한다.
            if candidate.count("�") > 10:
                continue
            text = candidate
            break
        except (LookupError, UnicodeDecodeError):
            continue
    if text is None:
        text = content.decode(encoding_hint or "utf-8", errors="replace")

    parser = "lxml"
    try:
        soup = BeautifulSoup(text, parser)
    except Exception:
        soup = BeautifulSoup(text, "html.parser")

    raw_tables: list[RawTable] = []
    for table_index, table in enumerate(soup.find_all("table")):
        cell_rows: list[list[CellSpec]] = []
        for tr in table.find_all("tr"):
            # 중첩 table의 tr을 상위 table 데이터로 잘못 포함하지 않는다.
            if tr.find_parent("table") is not table:
                continue
            cells = tr.find_all(["th", "td"], recursive=False)
            if not cells:
                continue
            row_specs: list[CellSpec] = []
            for cell in cells:
                row_specs.append(
                    CellSpec(
                        text=normalize_text(cell.get_text(" ", strip=True)),
                        is_header=(cell.name.lower() == "th"),
                        rowspan=int(cell.get("rowspan", 1) or 1),
                        colspan=int(cell.get("colspan", 1) or 1),
                    )
                )
            cell_rows.append(row_specs)
        if cell_rows:
            caption = normalize_text(table.get_text(" ", strip=True)[:500])
            raw_tables.append(
                RawTable(
                    cell_rows=cell_rows,
                    source=f"{source}#table-{table_index}",
                    caption=caption,
                )
            )
    return raw_tables


# ---------------------------------------------------------------------------
# requests 수집
# ---------------------------------------------------------------------------


def create_requests_session() -> Any:
    try:
        import requests
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
    except ImportError as exc:  # pragma: no cover - 환경 오류
        raise CollectorError(
            "requests가 설치되지 않았습니다. py -m pip install requests 를 실행하십시오."
        ) from exc

    session = requests.Session()
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    except Exception:
        pass
    retry = Retry(
        total=2,
        connect=2,
        read=2,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "POST"),
        raise_on_status=False,
    )
    session.mount("http://", HTTPAdapter(max_retries=retry))
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0 Safari/537.36"
            ),
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.7,en;q=0.6",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
    )
    return session


def collect_with_requests(
    session: Any, config: TerminalConfig, timeout_seconds: int
) -> ParsedTable:
    logging.info("[%s] requests 수집 시도: %s", config.code, config.url)
    response = session.get(
        config.url,
        timeout=(10, timeout_seconds),
        verify=False,
        allow_redirects=True,
    )
    if response.status_code >= 400:
        raise CollectorError(f"HTTP {response.status_code}: {response.url}")
    encoding_hint = response.encoding
    if not encoding_hint or encoding_hint.lower() in {"iso-8859-1", "ascii"}:
        encoding_hint = getattr(response, "apparent_encoding", None)
    raw_tables = raw_tables_from_html(response.content, encoding_hint, response.url)
    parsed = choose_best_table(raw_tables)
    if not parsed or len(parsed.rows) == 0:
        raise CollectorError("응답 HTML에서 선석배정 데이터 표를 찾지 못했습니다.")
    parsed.source = response.url
    return parsed


# ---------------------------------------------------------------------------
# Playwright 수집
# ---------------------------------------------------------------------------


TABLE_EXTRACT_SCRIPT = r"""
() => {
  const clean = (value) => (value || "")
    .replace(/\u200b/g, "")
    .replace(/[\t\r\n\u00a0 ]+/g, " ")
    .trim();

  return Array.from(document.querySelectorAll("table")).map((table, tableIndex) => ({
    tableIndex,
    caption: clean(table.innerText).slice(0, 500),
    rows: Array.from(table.rows).map((tr) =>
      Array.from(tr.cells).map((cell) => ({
        text: clean(cell.innerText),
        tag: cell.tagName.toLowerCase(),
        rowspan: Number(cell.getAttribute("rowspan") || 1),
        colspan: Number(cell.getAttribute("colspan") || 1)
      }))
    )
  }));
}
"""

DIV_ROWS_SCRIPT = r"""
(expectedColumns) => {
  const clean = (value) => (value || "")
    .replace(/\u200b/g, "")
    .replace(/[\t\r\n\u00a0 ]+/g, " ")
    .trim();
  const rows = [];
  const seen = new Set();
  for (const element of document.querySelectorAll("div")) {
    const children = Array.from(element.children).filter(
      (child) => child.tagName && child.tagName.toLowerCase() === "div"
    );
    if (children.length !== expectedColumns) continue;
    const values = children.map((child) => clean(child.innerText));
    const nonempty = values.filter(Boolean).length;
    if (nonempty < Math.max(2, Math.floor(expectedColumns / 5))) continue;
    const key = JSON.stringify(values);
    if (seen.has(key)) continue;
    seen.add(key);
    rows.push(values);
  }
  return rows;
}
"""

SCROLL_SCRIPT = r"""
() => {
  const elements = Array.from(document.querySelectorAll("body *"));
  const candidates = elements
    .filter((element) => {
      const style = getComputedStyle(element);
      const overflowY = style.overflowY;
      return (
        element.scrollHeight > element.clientHeight + 40 &&
        element.clientHeight > 80 &&
        (overflowY === "auto" || overflowY === "scroll" || overflowY === "overlay")
      );
    })
    .sort((a, b) => b.scrollHeight - a.scrollHeight)
    .slice(0, 12);

  let moved = 0;
  for (const element of candidates) {
    const before = element.scrollTop;
    const step = Math.max(120, Math.floor(element.clientHeight * 0.85));
    element.scrollTop = Math.min(element.scrollHeight, before + step);
    if (element.scrollTop > before) moved += 1;
  }
  const documentElement = document.scrollingElement || document.documentElement;
  const pageBefore = documentElement.scrollTop;
  documentElement.scrollTop = Math.min(
    documentElement.scrollHeight,
    pageBefore + Math.max(300, Math.floor(window.innerHeight * 0.85))
  );
  if (documentElement.scrollTop > pageBefore) moved += 1;
  return moved;
}
"""

SCROLL_TOP_SCRIPT = r"""
() => {
  for (const element of document.querySelectorAll("body *")) {
    if (element.scrollTop) element.scrollTop = 0;
  }
  const documentElement = document.scrollingElement || document.documentElement;
  documentElement.scrollTop = 0;
}
"""


class BrowserManager:
    def __init__(
        self,
        browser_name: str,
        headed: bool,
        timeout_seconds: int,
        debug_dir: Path,
        keep_debug: bool,
    ) -> None:
        self.browser_name = browser_name
        self.headed = headed
        self.timeout_ms = timeout_seconds * 1000
        self.debug_dir = debug_dir
        self.keep_debug = keep_debug
        self._playwright: Any = None
        self._browser: Any = None
        self.context: Any = None

    async def start(self) -> None:
        if self.context is not None:
            return
        try:
            from playwright.async_api import async_playwright
        except ImportError as exc:  # pragma: no cover - 환경 오류
            raise CollectorError(
                "Playwright 사전 환경 점검을 통과하지 못했습니다. "
                "terminal_schedule_collector.py --setup 을 실행하거나 "
                "자동 설치가 차단된 환경인지 확인하십시오."
            ) from exc

        self._playwright = await async_playwright().start()
        launch_options: dict[str, Any] = {
            "headless": not self.headed,
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-first-run",
                "--no-default-browser-check",
            ],
        }
        browser_type = self._playwright.chromium
        if self.browser_name in {"chrome", "msedge"}:
            launch_options["channel"] = self.browser_name

        try:
            self._browser = await browser_type.launch(**launch_options)
        except Exception as first_error:
            if "channel" in launch_options:
                logging.warning(
                    "설치된 %s 채널 실행 실패. Playwright Chromium으로 재시도합니다: %s",
                    self.browser_name,
                    sanitize_error(first_error),
                )
                launch_options.pop("channel", None)
                self._browser = await browser_type.launch(**launch_options)
            else:
                raise CollectorError(
                    "Playwright 브라우저 실행에 실패했습니다. "
                    "terminal_schedule_collector.py --setup 으로 환경을 다시 점검하십시오. "
                    f"원인: {sanitize_error(first_error)}"
                ) from first_error

        self.context = await self._browser.new_context(
            ignore_https_errors=True,
            locale="ko-KR",
            timezone_id="Asia/Seoul",
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0 Safari/537.36"
            ),
        )
        self.context.set_default_timeout(self.timeout_ms)
        self.context.set_default_navigation_timeout(self.timeout_ms)

    async def new_page(self) -> Any:
        await self.start()
        page = await self.context.new_page()
        page.on("dialog", lambda dialog: asyncio.create_task(dialog.dismiss()))
        return page

    async def capture_debug(self, page: Any, code: str, suffix: str) -> None:
        if not self.keep_debug:
            return
        self.debug_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(KST).strftime("%H%M%S")
        base = self.debug_dir / f"{code}_{timestamp}_{suffix}"
        try:
            await page.screenshot(path=str(base.with_suffix(".png")), full_page=True)
        except Exception:
            logging.debug("[%s] 디버그 스크린샷 저장 실패", code, exc_info=True)
        try:
            base.with_suffix(".html").write_text(await page.content(), encoding="utf-8")
        except Exception:
            logging.debug("[%s] 디버그 HTML 저장 실패", code, exc_info=True)

    async def close(self) -> None:
        if self.context is not None:
            await self.context.close()
            self.context = None
        if self._browser is not None:
            await self._browser.close()
            self._browser = None
        if self._playwright is not None:
            await self._playwright.stop()
            self._playwright = None


async def safe_wait(page: Any, milliseconds: int = 1800) -> None:
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=10_000)
    except Exception:
        pass
    await page.wait_for_timeout(milliseconds)


async def goto_page(page: Any, url: str, timeout_ms: int) -> None:
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
    except Exception as exc:
        # 일부 구형 HTTP 사이트는 load 이벤트가 끝나지 않지만 DOM은 이미 사용 가능하다.
        if "Timeout" not in type(exc).__name__ and "timeout" not in str(exc).lower():
            raise
        logging.warning("페이지 로드 제한시간 경과, 현재 DOM으로 계속 시도: %s", url)
    await safe_wait(page, 2200)


async def close_child_popups(context: Any, primary_page: Any) -> None:
    for page in list(context.pages):
        if page == primary_page:
            continue
        try:
            opener = await page.opener()
        except Exception:
            opener = None
        if opener == primary_page:
            try:
                await page.close()
            except Exception:
                pass


async def click_visible_locator(locator: Any) -> bool:
    try:
        count = await locator.count()
    except Exception:
        return False
    for index in range(min(count, 15)):
        item = locator.nth(index)
        try:
            if await item.is_visible() and await item.is_enabled():
                await item.click(timeout=5000)
                return True
        except Exception:
            continue
    return False


async def click_text_in_frames(page: Any, texts: Sequence[str]) -> bool:
    patterns = [re.compile(re.escape(text), re.IGNORECASE) for text in texts]
    for frame in page.frames:
        for pattern in patterns:
            try:
                if await click_visible_locator(frame.get_by_role("link", name=pattern)):
                    return True
            except Exception:
                pass
            try:
                if await click_visible_locator(frame.get_by_text(pattern)):
                    return True
            except Exception:
                pass
    # [2026-08-14] nexacro 메뉴 항목(div)은 액션체크에 걸려 일반 click이 스킵될 수 있다.
    # 텍스트 일치 요소를 force로 실제 클릭한다.
    for frame in page.frames:
        for pattern in patterns:
            try:
                loc = frame.get_by_text(pattern)
                count = await loc.count()
                for index in range(min(count, 6)):
                    try:
                        await loc.nth(index).click(force=True, timeout=4000)
                        return True
                    except Exception:
                        continue
            except Exception:
                pass
    return False


async def click_search_button(page: Any) -> bool:
    labels = ("조회", "검색")
    for frame in page.frames:
        for label in labels:
            pattern = re.compile(rf"^\s*{re.escape(label)}\s*$", re.IGNORECASE)
            for role in ("button", "link"):
                try:
                    if await click_visible_locator(frame.get_by_role(role, name=pattern)):
                        return True
                except Exception:
                    pass
            try:
                selector = (
                    f'input[type="button"][value="{label}"], '
                    f'input[type="submit"][value="{label}"], '
                    f'button:has-text("{label}")'
                )
                if await click_visible_locator(frame.locator(selector)):
                    return True
            except Exception:
                pass
            # [2026-08-14] nexacro 컴포넌트 버튼은 role=button이 아니라 텍스트를 가진 div로
            # 렌더되고, Playwright 액션체크(가시성/이벤트 수신)에 걸려 일반 click이 스킵될 수
            # 있다. 정확일치 조회 요소를 force로 실제 클릭한다(좌표 클릭과 동일한 신뢰 이벤트).
            try:
                exact = re.compile(rf"^\s*{re.escape(label)}\s*$")
                loc = frame.get_by_text(exact)
                count = await loc.count()
                for index in range(min(count, 8)):
                    try:
                        await loc.nth(index).click(force=True, timeout=4000)
                        return True
                    except Exception:
                        continue
            except Exception:
                pass
            # nexacro 조회 버튼 id에는 btn_find/btn_search/btn_inquiry 가 포함된다.
            try:
                loc = frame.locator('[id*="btn_find"], [id*="btn_search"], [id*="btn_inquiry"]')
                count = await loc.count()
                for index in range(min(count, 10)):
                    item = loc.nth(index)
                    try:
                        text = (await item.inner_text()).strip()
                    except Exception:
                        text = ""
                    if label in text or text == "":
                        try:
                            await item.click(force=True, timeout=4000)
                            return True
                        except Exception:
                            continue
            except Exception:
                pass
    return False


async def choose_week_period(page: Any) -> bool:
    for frame in page.frames:
        try:
            locator = frame.get_by_label(re.compile("일주일|1주"))
            if await click_visible_locator(locator):
                return True
        except Exception:
            pass
        try:
            labels = frame.locator("label")
            for index in range(min(await labels.count(), 40)):
                label = labels.nth(index)
                text = normalize_text(await label.inner_text())
                if "일주일" in text or text == "1주":
                    await label.click(timeout=3000)
                    return True
        except Exception:
            pass
        try:
            radios = frame.locator('input[type="radio"]')
            for index in range(min(await radios.count(), 30)):
                radio = radios.nth(index)
                value = normalize_text(await radio.get_attribute("value"))
                title = normalize_text(await radio.get_attribute("title"))
                name = normalize_text(await radio.get_attribute("name"))
                combined = f"{value} {title} {name}".lower()
                if any(token in combined for token in ("week", "7", "1w")):
                    await radio.check(force=True)
                    return True
        except Exception:
            pass
    return False


async def set_datatable_page_length(page: Any) -> None:
    for frame in page.frames:
        try:
            selects = frame.locator("select")
            count = await selects.count()
        except Exception:
            continue
        for index in range(min(count, 60)):
            select = selects.nth(index)
            try:
                name = normalize_text(await select.get_attribute("name")).lower()
                class_name = normalize_text(await select.get_attribute("class")).lower()
                if "length" not in name and "length" not in class_name:
                    continue
                options = await select.locator("option").all()
                candidates: list[tuple[int, str]] = []
                for option in options:
                    value = normalize_text(await option.get_attribute("value"))
                    text = normalize_text(await option.inner_text())
                    try:
                        number = int(value)
                    except ValueError:
                        try:
                            number = int(re.sub(r"\D", "", text))
                        except ValueError:
                            number = -1
                    candidates.append((10_000 if number == -1 else number, value))
                if candidates:
                    _, selected_value = max(candidates, key=lambda pair: pair[0])
                    await select.select_option(selected_value)
                    await frame.wait_for_timeout(700)
            except Exception:
                continue


async def collect_raw_tables_from_page(page: Any) -> list[RawTable]:
    raw_tables: list[RawTable] = []
    for frame_index, frame in enumerate(page.frames):
        try:
            descriptors = await frame.evaluate(TABLE_EXTRACT_SCRIPT)
        except Exception:
            continue
        for descriptor in descriptors or []:
            cell_rows: list[list[CellSpec]] = []
            for row in descriptor.get("rows", []):
                cell_rows.append(
                    [
                        CellSpec(
                            text=cell.get("text", ""),
                            is_header=(cell.get("tag") == "th"),
                            rowspan=int(cell.get("rowspan", 1) or 1),
                            colspan=int(cell.get("colspan", 1) or 1),
                        )
                        for cell in row
                    ]
                )
            if cell_rows:
                raw_tables.append(
                    RawTable(
                        cell_rows=cell_rows,
                        source=(
                            f"{frame.url or page.url}#frame-{frame_index}-"
                            f"table-{descriptor.get('tableIndex', 0)}"
                        ),
                        caption=descriptor.get("caption", ""),
                    )
                )
    return raw_tables


async def extract_best_browser_table(page: Any) -> ParsedTable | None:
    return choose_best_table(await collect_raw_tables_from_page(page))


async def generic_prepare(page: Any, config: TerminalConfig) -> None:
    await close_child_popups(page.context, page)
    try:
        await page.keyboard.press("Escape")
    except Exception:
        pass

    # 기본 조회 기간은 일주일을 우선 적용한다.
    await choose_week_period(page)

    if config.code == "DGT":
        # DGT는 기본 50건이므로 DataTables 페이지 길이를 최대로 변경한다.
        await set_datatable_page_length(page)

    # 결과가 자동 표시되지 않는 사이트를 위해 조회/검색 버튼을 한 번 누른다.
    clicked = await click_search_button(page)
    if clicked:
        await safe_wait(page, 2200)

    await set_datatable_page_length(page)
    await safe_wait(page, 900)


async def find_datatable_next(frame: Any) -> Any | None:
    selectors = (
        "a.paginate_button.next:not(.disabled)",
        "li.paginate_button.next:not(.disabled) a",
        "li.next:not(.disabled) a",
        "button.dt-paging-button.next:not(.disabled)",
        "button[aria-label='Next']:not([disabled])",
    )
    for selector in selectors:
        try:
            locator = frame.locator(selector)
            count = await locator.count()
            for index in range(min(count, 5)):
                item = locator.nth(index)
                if await item.is_visible() and await item.is_enabled():
                    aria_disabled = normalize_text(await item.get_attribute("aria-disabled")).lower()
                    class_name = normalize_text(await item.get_attribute("class")).lower()
                    if aria_disabled == "true" or "disabled" in class_name:
                        continue
                    return item
        except Exception:
            continue
    return None


async def collect_paginated_tables(
    page: Any,
    config: TerminalConfig,
    max_pages: int,
) -> ParsedTable:
    first = await extract_best_browser_table(page)
    if first is None:
        raise CollectorError("브라우저 DOM에서 선석배정 데이터 표를 찾지 못했습니다.")

    headers = first.headers
    all_rows: list[list[str]] = list(first.rows)
    seen_pages: set[tuple[tuple[str, ...], ...]] = {
        tuple(tuple(row) for row in first.rows[:5])
    }

    # iCON은 지침에 따라 1페이지만 수집한다.
    if config.code == "ICON":
        return first

    for _ in range(1, max_pages):
        next_locator = None
        for frame in page.frames:
            next_locator = await find_datatable_next(frame)
            if next_locator is not None:
                break
        if next_locator is None:
            break

        try:
            await next_locator.click(timeout=5000)
            await safe_wait(page, 900)
        except Exception:
            break

        current = await extract_best_browser_table(page)
        if current is None:
            break
        signature = tuple(tuple(row) for row in current.rows[:5])
        if signature in seen_pages:
            break
        seen_pages.add(signature)

        if len(current.headers) != len(headers):
            logging.warning(
                "[%s] 페이지 이동 후 컬럼 수 변경(%d -> %d), 추가 페이지 중단",
                config.code,
                len(headers),
                len(current.headers),
            )
            break
        all_rows.extend(current.rows)

    return ParsedTable(
        headers=headers,
        rows=deduplicate_rows(all_rows),
        score=first.score,
        source=first.source,
        diagnostics=first.diagnostics,
    )


# ---------------------------------------------------------------------------
# nexacro(TOBESOFT) SPA 수집: DOM/클릭이 아니라 애플리케이션 JS API로 직접
# 폼을 열고 조회(fn_find)한 뒤 결과 데이터셋(ds_list)을 읽는다. (2026-08-14)
# ---------------------------------------------------------------------------
NEXACRO_SCRIPT = r"""async (cfg) => {
  const sleep = ms => new Promise(r => setTimeout(r, ms));
  const app = window.application;
  for (let i = 0; i < 30 && !(app && app.mainframe && app.mainframe.vframeset); i++) await sleep(500);
  if (!app || !app.mainframe) return { error: "no nexacro application" };
  const kids = o => { const f = o && (o.frames || o._frames); const a = []; if (f) { const n = f.length || 0; for (let i = 0; i < n; i++) if (f[i]) a.push(f[i]); } return a; };
  try { const pfs = app.popupframes; if (pfs) for (let i = pfs.length - 1; i >= 0; i--) { try { if (pfs[i] && pfs[i].form && pfs[i].form.close) pfs[i].form.close(); } catch (e) {} } } catch (e) {}
  await sleep(500);
  let lf; try { lf = app.mainframe.vframeset.hframeset.leftframe.form; } catch (e) { return { error: "no leftframe" }; }
  const dm = lf.ds_menu; const key = cfg.menuName.replace(/\s+/g, "");
  let mi = -1; for (let r = 0; r < dm.getRowCount(); r++) { if ((dm.getColumn(r, "menuNm") || "").replace(/\s+/g, "").indexOf(key) >= 0) { mi = r; break; } }
  if (mi < 0) return { error: "menu not found: " + cfg.menuName };
  const menuId = dm.getColumn(mi, "menuId"), menuNm = dm.getColumn(mi, "menuNm"), pgmUrl = dm.getColumn(mi, "pgmUrl");
  dm.set_rowposition(mi);
  try { lf.gfn_openMenu(menuId, menuNm, pgmUrl); } catch (e) { return { error: "openMenu: " + e.message }; }
  let dw = null;
  for (let t = 0; t < 30 && !dw; t++) { await sleep(500);
    try { const hf = app.mainframe.vframeset.hframeset; const bf = kids(hf).find(f => /body/i.test(f.name || "")); const wf = kids(bf).find(f => /work/i.test(f.name || "")); const cf = kids(wf).find(f => (f.name || "") === menuId); if (cf && cf.form && cf.form.div_commBtn && cf.form.div_commBtn.form) dw = cf.form.div_commBtn.form; } catch (e) {}
  }
  if (!dw) return { error: "work form not found: " + menuId };
  if (cfg.setDate) { try { const c = dw.ds_cond; const d = new Date(); const p = x => ("" + x).padStart(2, "0"); const t = d.getFullYear() + p(d.getMonth() + 1) + p(d.getDate()); const e2 = new Date(d.getTime() + 7 * 864e5); const en = e2.getFullYear() + p(e2.getMonth() + 1) + p(e2.getDate()); for (let i = 0; i < c.getColCount(); i++) { const id = c.getColID(i); if (/frdate|fromdate|fr_date/i.test(id)) c.setColumn(0, id, t); if (/todate|to_date|enddate/i.test(id)) c.setColumn(0, id, en); } } catch (e) {} }
  try { if (typeof dw.fn_find === "function") dw.fn_find(); } catch (e) {}
  const ds = dw.ds_list; let n = 0; for (let t = 0; t < 24; t++) { await sleep(500); n = ds.getRowCount(); if (n > 0) break; }
  const clean = v => ("" + (v == null ? "" : v)).replace(/\s+/g, " ").trim();
  const rows = []; for (let r = 0; r < n; r++) { rows.push(cfg.dsCols.map(c => c === "#ROWNUM" ? String(r + 1) : clean(ds.getColumn(r, c)))); }
  return { rowCount: n, rows: rows };
}"""

NEXACRO_CONFIG = {
    "PNCT": {"menuName": "선석 스케쥴(텍스트)", "setDate": False,
             "headers": ["선석", "선사", "모선항차(선사항차)", "선명", "Route", "반입마감시한", "접안(예정)일시", "출항(예정)일시", "상태"],
             "dsCols": ["BERTH", "OPERATOR", "VVD_BITT", "VSL_NAME", "ROUTE", "CUT_OFF_DATE", "ETB_DATE", "ETD_DATE", "STATUS"]},
    "DDCT": {"menuName": "선석 배정 현황(텍스트)", "setDate": True,
             "headers": ["No", "모선/항차", "선사", "입항", "출항", "CCT", "ETB/ATB", "ETD/ATD", "양하", "적하", "이적", "선석", "모선명"],
             "dsCols": ["#ROWNUM", "plvVslvoy", "cdvOperator", "plvEvoyin", "plvEvoyout", "cct", "plvAtb", "plvAtd", "plvDisvan", "plvLodvan", "plvShiftvan", "plvBerth", "cdvName"]},
    "BCT": {"menuName": "선석 배정현황 (목록)", "setDate": True,
            "headers": ["No", "선석", "선사", "모선/항차", "입항", "출항", "CCT", "접안예정(ETB)", "출항예정(ETD)", "양하", "적하", "이적", "모선명", "ROUTE", "전배TML", "검역", "줄잡이업체", "상태"],
            "dsCols": ["#ROWNUM", "plvBerth", "cdvOperator", "plvVslvoy", "plvEvoyin", "plvEvoyout", "cct", "plvAtb", "plvAtd", "plvDisvan", "plvLodvan", "plvShiftvan", "cdvName", "plvRoute", "plvNeartml", "plvQuarantine", "lineCop", "plvStatus"]},
}


async def collect_nexacro(browser: BrowserManager, config: TerminalConfig) -> ParsedTable:
    cfg = NEXACRO_CONFIG[config.code]
    page = await browser.new_page()
    try:
        logging.info("[%s] nexacro API 수집: %s", config.code, config.url)
        await goto_page(page, config.url, browser.timeout_ms)
        await page.wait_for_timeout(3000)
        result = await page.evaluate(NEXACRO_SCRIPT, cfg)
        if not result or result.get("error"):
            raise CollectorError("nexacro 수집 실패: %s" % ((result or {}).get("error", "no result")))
        raw_rows = result.get("rows") or []
        if not raw_rows:
            raise CollectorError("nexacro ds_list가 비어 있습니다.")
        headers = list(cfg["headers"])
        width = len(headers)
        data = []
        for row in raw_rows:
            row = [normalize_text(v) for v in row]
            if len(row) < width:
                row = row + [""] * (width - len(row))
            data.append(row[:width])
        return ParsedTable(headers=headers, rows=data, score=100 + len(data), source=page.url)
    except Exception:
        await browser.capture_debug(page, config.code, "nexacro_failed")
        raise
    finally:
        await page.close()


async def collect_generic_browser(
    browser: BrowserManager,
    config: TerminalConfig,
    max_pages: int,
) -> ParsedTable:
    page = await browser.new_page()
    try:
        logging.info("[%s] Playwright 수집 시도: %s", config.code, config.url)
        await goto_page(page, config.url, browser.timeout_ms)
        await generic_prepare(page, config)
        parsed = await collect_paginated_tables(page, config, max_pages=max_pages)
        parsed.source = page.url
        return parsed
    except Exception:
        await browser.capture_debug(page, config.code, "generic_failed")
        raise
    finally:
        await page.close()


async def collect_legacy_hutchison(
    browser: BrowserManager,
    config: TerminalConfig,
    max_pages: int,
) -> ParsedTable:
    page = await browser.new_page()
    try:
        logging.info("[%s] 레거시 페이지 전량 수집: %s", config.code, config.url)
        await goto_page(page, config.url, browser.timeout_ms)
        await safe_wait(page, 1200)

        first = await extract_best_browser_table(page)
        if first is None:
            raise CollectorError("Closing Time이 포함된 선석배정 표를 찾지 못했습니다.")

        headers = first.headers
        all_rows = list(first.rows)
        signatures: set[tuple[tuple[str, ...], ...]] = {
            tuple(tuple(row) for row in first.rows[:5])
        }

        calls: list[tuple[Any, str]] = []
        for frame in page.frames:
            try:
                hrefs = await frame.locator("a[href*='gotoPage']").evaluate_all(
                    "els => els.map(e => e.getAttribute('href')).filter(Boolean)"
                )
            except Exception:
                continue
            for href in hrefs:
                code = re.sub(r"^javascript:\s*", "", str(href), flags=re.I).strip()
                if code and all(existing != code for _, existing in calls):
                    calls.append((frame, code))

        # 실제 페이지 링크가 보이면 해당 링크만 사용한다. 링크가 없을 때만 알려진 함수 규격을 보수적으로 시도한다.
        if not calls:
            frame = page.main_frame
            for page_no in range(2, max_pages + 1):
                calls.append((frame, f"gotoPage('1','{page_no}','T01','01')"))

        for frame, call in calls[: max(0, max_pages - 1)]:
            try:
                await frame.evaluate("code => eval(code)", call)
                await safe_wait(page, 800)
            except Exception:
                logging.debug("[%s] 페이지 함수 실행 실패: %s", config.code, call)
                continue

            current = await extract_best_browser_table(page)
            if current is None:
                continue
            signature = tuple(tuple(row) for row in current.rows[:5])
            if signature in signatures:
                continue
            signatures.add(signature)
            if len(current.headers) != len(headers):
                logging.warning(
                    "[%s] 페이지별 컬럼 수 불일치(%d != %d), 해당 페이지 제외",
                    config.code,
                    len(current.headers),
                    len(headers),
                )
                continue
            all_rows.extend(current.rows)

        return ParsedTable(
            headers=headers,
            rows=deduplicate_rows(all_rows),
            score=first.score,
            source=page.url,
            diagnostics=first.diagnostics,
        )
    except Exception:
        await browser.capture_debug(page, config.code, "legacy_failed")
        raise
    finally:
        await page.close()


async def collect_div_rows_once(page: Any, expected_columns: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for frame in page.frames:
        try:
            frame_rows = await frame.evaluate(DIV_ROWS_SCRIPT, expected_columns)
        except Exception:
            continue
        rows.extend([[normalize_text(v) for v in row] for row in (frame_rows or [])])
    return deduplicate_rows(rows)


async def collect_virtual_div_rows(
    page: Any,
    expected_columns: int,
    max_steps: int = 30,
) -> list[list[str]]:
    collected: list[list[str]] = []
    seen: set[tuple[str, ...]] = set()
    stagnant = 0

    for _ in range(max_steps):
        new_count = 0
        for row in await collect_div_rows_once(page, expected_columns):
            key = tuple(row)
            if key not in seen:
                seen.add(key)
                collected.append(row)
                new_count += 1

        moved = 0
        for frame in page.frames:
            try:
                moved += int(await frame.evaluate(SCROLL_SCRIPT) or 0)
            except Exception:
                continue
        await page.wait_for_timeout(350)

        if new_count == 0:
            stagnant += 1
        else:
            stagnant = 0
        if moved == 0 or stagnant >= 4:
            break

    # 마지막 위치에서 한 번 더 수집한다.
    for row in await collect_div_rows_once(page, expected_columns):
        key = tuple(row)
        if key not in seen:
            seen.add(key)
            collected.append(row)

    for frame in page.frames:
        try:
            await frame.evaluate(SCROLL_TOP_SCRIPT)
        except Exception:
            pass
    return collected


def parsed_table_from_div_rows(rows: Sequence[Sequence[str]], source: str) -> ParsedTable:
    if len(rows) < 2:
        raise CollectorError("div 기반 그리드 행을 충분히 찾지 못했습니다.")
    width = len(rows[0])
    normalized = [list(map(normalize_text, row)) for row in rows if len(row) == width]
    if len(normalized) < 2:
        raise CollectorError("div 기반 그리드의 컬럼 수가 일정하지 않습니다.")

    header_index = max(
        range(min(len(normalized), 15)),
        key=lambda index: row_header_score(normalized[index]),
    )
    if row_header_score(normalized[header_index]) < 8:
        raise CollectorError("div 기반 그리드에서 원본 헤더 행을 식별하지 못했습니다.")

    headers = unique_headers(normalized[header_index])
    data_rows: list[list[str]] = []
    for row in normalized:
        if row == normalized[header_index] or is_footer_or_noise(row):
            continue
        if sum(bool(v) for v in row) < max(2, width // 5):
            continue
        # 데이터 행은 날짜 또는 다수의 숫자/문자 필드를 가져야 한다.
        joined = " ".join(row)
        value_density = sum(bool(v) for v in row) / width
        if not DATE_PATTERN.search(joined) and value_density < 0.45:
            continue
        data_rows.append(row)
    data_rows = deduplicate_rows(data_rows)
    if not data_rows:
        raise CollectorError("div 기반 그리드에서 데이터 행을 찾지 못했습니다.")

    return ParsedTable(
        headers=headers,
        rows=data_rows,
        score=100 + len(data_rows),
        source=source,
        diagnostics={"width": width, "header_index": header_index},
    )


async def dismiss_visible_modal_close_buttons(page: Any) -> None:
    close_patterns = (
        re.compile(r"^\s*닫기\s*$"),
        re.compile(r"^\s*오늘\s*하루.*열지.*$"),
        re.compile(r"^\s*[xX]\s*$"),
    )
    for frame in page.frames:
        for pattern in close_patterns:
            try:
                locator = frame.get_by_role("button", name=pattern)
                await click_visible_locator(locator)
            except Exception:
                pass
            try:
                locator = frame.get_by_text(pattern)
                await click_visible_locator(locator)
            except Exception:
                pass
    # [2026-08-14] nexacro 팝업 닫기(X) 버튼은 id에 btn_close 를 포함(텍스트 'X').
    # Playwright 액션체크에 걸리므로 force로 실제 클릭해 공지 팝업을 닫는다.
    for frame in page.frames:
        try:
            loc = frame.locator('[id*="btn_close"]')
            count = await loc.count()
            for index in range(min(count, 8)):
                try:
                    await loc.nth(index).click(force=True, timeout=3000)
                except Exception:
                    continue
        except Exception:
            pass


async def collect_div_grid(
    browser: BrowserManager,
    config: TerminalConfig,
) -> ParsedTable:
    assert config.expected_columns is not None
    page = await browser.new_page()
    try:
        logging.info("[%s] div 기반 그리드 수집: %s", config.code, config.url)
        await goto_page(page, config.url, browser.timeout_ms)
        await close_child_popups(page.context, page)
        await dismiss_visible_modal_close_buttons(page)

        if config.menu_texts:
            clicked = await click_text_in_frames(page, config.menu_texts)
            if not clicked:
                logging.warning("[%s] 메뉴 텍스트 클릭 실패, 현재 화면에서 계속 시도", config.code)
            await safe_wait(page, 1800)
            await dismiss_visible_modal_close_buttons(page)

        await choose_week_period(page)
        if await click_search_button(page):
            await safe_wait(page, 1800)

        # [2026-08-14] 메뉴/조회 후 그리드가 비동기로 채워질 때까지 최대 약 8초 폴링.
        rows = []
        for _ in range(8):
            await safe_wait(page, 1000)
            rows = await collect_virtual_div_rows(page, config.expected_columns)
            if len(rows) >= 2:
                break
        return parsed_table_from_div_rows(rows, page.url)
    except Exception:
        await browser.capture_debug(page, config.code, "div_grid_failed")
        raise
    finally:
        await page.close()


async def find_menu_anchor(page: Any, texts: Sequence[str]) -> tuple[Any, str | None] | None:
    for frame in page.frames:
        try:
            anchors = frame.locator("a")
            count = await anchors.count()
        except Exception:
            continue
        for index in range(min(count, 300)):
            anchor = anchors.nth(index)
            try:
                text = normalize_text(await anchor.inner_text())
            except Exception:
                continue
            compact = re.sub(r"\s+", "", text)
            if not any(re.sub(r"\s+", "", target) in compact for target in texts):
                continue
            try:
                if not await anchor.is_visible():
                    continue
            except Exception:
                pass
            return anchor, await anchor.get_attribute("href")
    return None


async def collect_bpt(
    browser: BrowserManager,
    config: TerminalConfig,
    max_pages: int,
) -> ParsedTable:
    home = await browser.new_page()
    target_page = home
    try:
        logging.info("[%s] BPT 전용 수집: %s", config.code, config.url)
        await goto_page(home, config.url, browser.timeout_ms)  # 홈 진입(세션 쿠키 확보)
        # [2026-08-14] 퀵메뉴 '선석배정 현황(T)'가 새 탭으로 여는 실제 URL로 직접 진입한다.
        # 홈의 동일 텍스트 <a>는 '#' 앵커이거나 .jsp 없는 404 링크라 페이지 진입에
        # 실패하므로, find_menu_anchor 대신 확정된 berth_status URL로 곧장 이동한다.
        berth_url = urljoin(
            home.url,
            "/content/sw/frame/berth_status_text_frame_sw_kr.jsp"
            "?p_id=CONT_CN_KR&search=Y&snb_num=2",
        )
        await goto_page(home, berth_url, browser.timeout_ms)
        target_page = home
        await safe_wait(target_page, 1500)

        await choose_week_period(target_page)
        if await click_search_button(target_page):
            await safe_wait(target_page, 2200)
        await set_datatable_page_length(target_page)
        parsed = await collect_paginated_tables(target_page, config, max_pages=max_pages)
        parsed.source = target_page.url
        return parsed
    except Exception:
        await browser.capture_debug(target_page, config.code, "bpt_failed")
        raise
    finally:
        if target_page != home:
            try:
                await target_page.close()
            except Exception:
                pass
        await home.close()


async def collect_dongbang_grid(
    browser: BrowserManager,
    config: TerminalConfig,
) -> ParsedTable:
    assert config.expected_columns is not None
    page = await browser.new_page()
    try:
        logging.info("[%s] 동방 infoservice 수집: %s", config.code, config.url)
        await goto_page(page, config.url, browser.timeout_ms)
        await safe_wait(page, 1000)
        await close_child_popups(page.context, page)
        await dismiss_visible_modal_close_buttons(page)

        clicked = await click_text_in_frames(page, config.menu_texts)
        if not clicked:
            raise CollectorError("좌측 선석배정 텍스트 메뉴를 찾지 못했습니다.")
        await safe_wait(page, 1800)
        await dismiss_visible_modal_close_buttons(page)

        if config.code == "DDCT":
            await choose_week_period(page)
        await click_search_button(page)

        # [2026-08-14] nexacro 그리드는 조회 후 데이터가 비동기로 채워진다.
        # 행이 채워질 때까지 최대 약 8초간 폴링한 뒤 추출한다.
        rows: list[list[str]] = []
        for _ in range(8):
            await safe_wait(page, 1000)
            table = await extract_best_browser_table(page)
            if table and (
                config.expected_columns is None
                or len(table.headers) == config.expected_columns
            ):
                table.source = page.url
                return table
            rows = await collect_virtual_div_rows(page, config.expected_columns)
            if len(rows) >= 2:
                break
        return parsed_table_from_div_rows(rows, page.url)
    except Exception:
        await browser.capture_debug(page, config.code, "dongbang_failed")
        raise
    finally:
        await page.close()


# ---------------------------------------------------------------------------
# 결과 검증
# ---------------------------------------------------------------------------


def validate_parsed_table(config: TerminalConfig, parsed: ParsedTable) -> ParsedTable:
    headers = unique_headers(parsed.headers)
    width = len(headers)
    if width < 4:
        raise CollectorError(f"헤더 컬럼 수가 너무 적습니다: {width}")
    if config.expected_columns is not None and width != config.expected_columns:
        raise CollectorError(
            f"컬럼 수 검증 실패: 사이트 헤더 {width}개, 예상 {config.expected_columns}개"
        )

    normalized_rows: list[list[str]] = []
    mismatches: list[tuple[int, int]] = []
    for index, row in enumerate(parsed.rows, start=1):
        normalized = [normalize_text(v) for v in row]
        if len(normalized) != width:
            mismatches.append((index, len(normalized)))
            continue
        normalized_rows.append(normalized)

    if mismatches:
        preview = ", ".join(f"{row_no}행={count}개" for row_no, count in mismatches[:10])
        raise CollectorError(
            f"행별 컬럼 정렬 검증 실패: 헤더 {width}개, {preview}"
        )
    normalized_rows = deduplicate_rows(normalized_rows)
    if not normalized_rows:
        raise CollectorError("검증 후 유효 데이터 행이 없습니다.")

    parsed.headers = headers
    parsed.rows = normalized_rows
    return parsed


# ---------------------------------------------------------------------------
# Excel / HTML / JSON 생성
# ---------------------------------------------------------------------------


def import_openpyxl() -> Any:
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:  # pragma: no cover - 환경 오류
        raise CollectorError(
            "openpyxl이 설치되지 않았습니다. py -m pip install openpyxl 를 실행하십시오."
        ) from exc


def populate_worksheet(ws: Any, result: TerminalResult) -> None:
    """첨부 샘플 통합 엑셀과 동일한 형식으로 한 터미널 시트를 작성한다."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # 샘플 기준: 회색 헤더, Arial 10pt, #999999 얇은 테두리.
    thin = Side(style="thin", color=EXCEL_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor=EXCEL_HEADER_FILL)
    body_font = Font(name="Arial", size=EXCEL_BODY_FONT_SIZE)
    header_font = Font(name="Arial", size=EXCEL_BODY_FONT_SIZE, bold=True)
    source_font = Font(
        name="Arial",
        size=EXCEL_SOURCE_FONT_SIZE,
        italic=True,
        color=EXCEL_SOURCE_FONT_COLOR,
    )
    body_alignment = Alignment(vertical="center")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [safe_excel_text(value) for value in result.headers]
    rows = [[safe_excel_text(value) for value in row] for row in result.rows]

    ws.append(headers)
    for row in rows:
        ws.append(row)

    max_row = 1 + len(rows)
    max_col = len(headers)

    # 샘플은 첫 행을 회색 굵은 헤더로 표시하며 줄바꿈은 사용하지 않는다.
    # 헤더를 먼저 적용해 샘플 파일의 스타일 등록 순서까지 동일하게 유지한다.
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = header_alignment

    # 데이터 본문은 Arial 10pt, 세로 가운데, #999999 얇은 테두리다.
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = body_alignment

    # 샘플과 동일하게 첫 행만 고정하고, 자동 필터/격자선 숨김은 적용하지 않는다.
    ws.freeze_panes = "A2"

    # 한글은 2칸, 영문/숫자는 1칸으로 계산한 최대 표시폭 + 2. 최소 너비는 6.
    # 출처 문구는 열 너비 계산에서 제외한다.
    for column_index in range(1, max_col + 1):
        candidates = [headers[column_index - 1]] + [
            row[column_index - 1] for row in rows
        ]
        width = max(
            (display_width(str(value)) for value in candidates),
            default=EXCEL_MIN_COLUMN_WIDTH - 2,
        ) + 2
        ws.column_dimensions[get_column_letter(column_index)].width = max(
            width, EXCEL_MIN_COLUMN_WIDTH
        )

    # 데이터 마지막 행 아래 한 줄을 비우고 출처/수집시각을 기록한다.
    source_row = max_row + 2
    source_text = (
        f"출처: {result.source_url or result.config.url}, "
        f"수집시각: {result.collected_at.strftime('%Y-%m-%d %H:%M')} (KST)"
    )
    ws.cell(source_row, 1, source_text)
    ws.cell(source_row, 1).font = source_font


def verify_excel_sample_format(
    workbook_path: Path,
    sheet_name: str,
    expected_headers: Sequence[str],
    expected_rows: Sequence[Sequence[str]],
) -> None:
    """자체 점검에서 첨부 샘플의 엑셀 출력 형식이 유지되는지 검증한다."""
    from openpyxl.utils import get_column_letter

    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(workbook_path)
    try:
        worksheet = workbook[sheet_name]
        data_last_row = 1 + len(expected_rows)
        source_row = data_last_row + 2

        header = worksheet.cell(1, 1)
        body = worksheet.cell(2, 1) if expected_rows else None
        source = worksheet.cell(source_row, 1)

        def rgb_suffix(color: Any) -> str:
            rgb = getattr(color, "rgb", None)
            return str(rgb or "")[-6:].upper()

        assert header.font.name == "Arial"
        assert float(header.font.sz or 0) == float(EXCEL_BODY_FONT_SIZE)
        assert header.font.bold is True
        assert rgb_suffix(header.fill.fgColor) == EXCEL_HEADER_FILL
        assert header.alignment.horizontal == "center"
        assert header.alignment.vertical == "center"
        assert header.alignment.wrap_text is not True

        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(header.border, side_name)
            assert side.style == "thin"
            assert rgb_suffix(side.color) == EXCEL_BORDER_COLOR

        if body is not None:
            assert body.font.name == "Arial"
            assert float(body.font.sz or 0) == float(EXCEL_BODY_FONT_SIZE)
            assert body.font.bold is not True
            assert body.alignment.vertical == "center"
            assert body.alignment.horizontal is None
            for side_name in ("left", "right", "top", "bottom"):
                side = getattr(body.border, side_name)
                assert side.style == "thin"
                assert rgb_suffix(side.color) == EXCEL_BORDER_COLOR

        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref is None
        assert worksheet.sheet_view.showGridLines is not False
        assert not worksheet.merged_cells.ranges
        assert len(worksheet.tables) == 0
        assert worksheet.cell(source_row - 1, 1).value is None
        assert source.value and str(source.value).startswith("출처: ")
        assert source.font.name == "Arial"
        assert float(source.font.sz or 0) == float(EXCEL_SOURCE_FONT_SIZE)
        assert source.font.italic is True
        assert rgb_suffix(source.font.color) == EXCEL_SOURCE_FONT_COLOR
        assert source.border.left.style is None

        safe_headers = [safe_excel_text(value) for value in expected_headers]
        safe_rows = [
            [safe_excel_text(value) for value in row] for row in expected_rows
        ]
        for column_index in range(1, len(safe_headers) + 1):
            candidates = [safe_headers[column_index - 1]] + [
                row[column_index - 1] for row in safe_rows
            ]
            expected_width = max(
                max(display_width(str(value)) for value in candidates) + 2,
                EXCEL_MIN_COLUMN_WIDTH,
            )
            actual_width = worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width
            assert float(actual_width or 0) == float(expected_width)
    finally:
        workbook.close()


def save_terminal_workbook(result: TerminalResult, root: Path, run_date: date) -> Path:
    openpyxl = import_openpyxl()
    folder = root / result.config.folder
    folder.mkdir(parents=True, exist_ok=True)
    output = folder / f"{result.config.code}_선석배정현황_{run_date:%Y%m%d}.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = result.config.code
    populate_worksheet(worksheet, result)
    workbook.save(output)
    return output


def save_combined_workbook(
    results: Sequence[TerminalResult], root: Path, run_date: date
) -> Path:
    openpyxl = import_openpyxl()
    successful = {result.config.code: result for result in results if result.success}
    if not successful:
        raise CollectorError("통합 Excel을 생성할 성공 터미널이 없습니다.")

    target_dir = root / "통합" / f"{run_date:%Y}" / f"{run_date:%m}"
    target_dir.mkdir(parents=True, exist_ok=True)
    output = target_dir / f"터미널_선석배정현황_통합_{run_date:%Y%m%d}.xlsx"

    if output.exists():
        # [2026-08-14] 병합 저장: 기존 통합엑셀을 열어 이번에 수집 성공한 시트만 교체한다.
        # 수집 실패한 터미널(예: BPT/BCT/PNCT/DDCT)은 기존(앞선 수집분) 데이터를 그대로
        # 보존하여, 폴백 수집이 더 완전한 수집분을 덮어써 데이터가 줄어드는 것을 방지한다.
        workbook = openpyxl.load_workbook(output)
        for code in SHEET_ORDER:
            result = successful.get(code)
            if result is None:
                continue
            if code in workbook.sheetnames:
                index = workbook.sheetnames.index(code)
                workbook.remove(workbook[code])
                worksheet = workbook.create_sheet(code, index)
            else:
                worksheet = workbook.create_sheet(code)
            populate_worksheet(worksheet, result)
    else:
        workbook = openpyxl.Workbook()
        default = workbook.active
        workbook.remove(default)
        for code in SHEET_ORDER:
            result = successful.get(code)
            if result is None:
                continue
            worksheet = workbook.create_sheet(code)
            populate_worksheet(worksheet, result)
    order_index = {code: pos for pos, code in enumerate(SHEET_ORDER)}
    workbook._sheets.sort(key=lambda sheet: order_index.get(sheet.title, len(order_index)))
    workbook.save(output)
    return output


def html_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    parts = [
        '<table border="1" cellspacing="0" cellpadding="3" '
        'style="border-collapse:collapse;border-color:#999;font-size:12px">',
        '<tr style="background-color:#D9D9D9;font-weight:bold">',
    ]
    for header in headers:
        parts.append(f"<th>{html_lib.escape(normalize_text(header))}</th>")
    parts.append("</tr>")
    for row in rows:
        parts.append("<tr>")
        for value in row:
            parts.append(f"<td>{html_lib.escape(normalize_text(value))}</td>")
        parts.append("</tr>")
    parts.append("</table>")
    return "".join(parts)


def create_report_html(
    results: Sequence[TerminalResult],
    run_date: date,
    output_path: Path,
) -> Path:
    total = sum(result.count for result in results if result.success)
    port_order = ["부산신항", "부산북항", "광양항", "인천항", "평택당진항", "대산항"]
    grouped: dict[str, list[TerminalResult]] = defaultdict(list)
    for result in results:
        grouped[result.config.port].append(result)

    body: list[str] = [
        "<!doctype html>",
        '<html lang="ko"><head><meta charset="utf-8">',
        "<title>국내 컨테이너 터미널 선석배정현황 일일 보고</title>",
        "</head><body style=\"font-family:'Malgun Gothic',Arial,sans-serif;font-size:13px;color:#222\">",
        "<p>안녕하세요.</p>",
        "<p>국내 주요 컨테이너 터미널의 선석배정현황을 아래와 같이 보고드립니다.</p>",
        "<h3>[1] 수집 개요</h3>",
        "<ul>",
        f"<li>기준일자: {run_date:%Y-%m-%d}</li>",
        "<li>수집대상: 부산신항, 부산북항, 광양항, 인천항, 평택당진항, 대산항 16개 터미널</li>",
        f"<li>총 선석배정 건수: {total:,}건</li>",
        "</ul>",
        "<h3>[2] 터미널별 수집 건수 요약</h3>",
    ]

    summary_headers = ["항만", "코드", "터미널", "상태", "건수", "비고"]
    summary_rows: list[list[str]] = []
    for port in port_order:
        for result in grouped.get(port, []):
            status = "성공" if result.success else "실패"
            note = result.note if result.success else result.error
            summary_rows.append(
                [
                    port,
                    result.config.code,
                    result.config.name,
                    status,
                    f"{result.count:,}" if result.success else "-",
                    note,
                ]
            )
    body.append(html_table(summary_headers, summary_rows))

    body.append("<h3>[3] 터미널별 상세내역</h3>")
    for port in port_order:
        for result in grouped.get(port, []):
            title = (
                f"{result.config.port} - {result.config.name} "
                f"({result.config.code}, {result.count:,}건)"
            )
            body.append(f"<h4>{html_lib.escape(title)}</h4>")
            if result.success:
                if result.note:
                    body.append(f"<p>{html_lib.escape(result.note)}</p>")
                body.append(html_table(result.headers, result.rows))
            else:
                body.append(
                    "<p style=\"color:#9C0006\">수집 실패: "
                    f"{html_lib.escape(result.error)}</p>"
                )

    body.extend(
        [
            "<h3>[4] 첨부자료</h3>",
            "<p>통합 엑셀 파일(터미널별 시트)을 첨부하였습니다.</p>",
            "<p>본 보고서는 매일 오전 자동 수집되어 발송됩니다.</p>",
            "<p>감사합니다.</p>",
            "</body></html>",
        ]
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(body), encoding="utf-8")
    return output_path


def create_pending_json(
    root: Path,
    windows_root: PureWindowsPath,
    run_date: date,
    html_path: Path,
    combined_path: Path,
) -> Path:
    pending_dir = root / "_tools" / "pending"
    pending_dir.mkdir(parents=True, exist_ok=True)
    output = pending_dir / f"report_{run_date:%Y%m%d}.json"

    html_relative = html_path.relative_to(root)
    combined_relative = combined_path.relative_to(root)
    subject = f"국내 컨테이너 터미널 선석배정현황 일일 보고 ({run_date:%Y-%m-%d})"
    # Outlook COM 경로에서 CP949로 안전하게 변환 가능한지 사전 검증한다.
    subject.encode("cp949", errors="strict")

    payload = {
        "to": ALLOWED_RECIPIENT,
        "subject": subject,
        "html": to_windows_path(windows_root, html_relative),
        "attach": [to_windows_path(windows_root, combined_relative)],
        "cc": "",
    }
    output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output


def load_previous_history(history_dir: Path, run_date: date) -> dict[str, Any] | None:
    candidates: list[tuple[date, Path]] = []
    for path in history_dir.glob("run_*.json"):
        match = re.fullmatch(r"run_(\d{8})\.json", path.name)
        if not match:
            continue
        try:
            file_date = datetime.strptime(match.group(1), "%Y%m%d").date()
        except ValueError:
            continue
        if file_date < run_date:
            candidates.append((file_date, path))
    if not candidates:
        return None
    _, latest = max(candidates, key=lambda pair: pair[0])
    try:
        return json.loads(latest.read_text(encoding="utf-8"))
    except Exception:
        return None


def create_history_json(
    results: Sequence[TerminalResult],
    run_date: date,
    history_dir: Path,
    combined_path: Path | None,
    html_path: Path,
    pending_path: Path | None,
) -> tuple[Path, str]:
    previous = load_previous_history(history_dir, run_date)
    counts = {
        result.config.code: result.count if result.success else None for result in results
    }
    comparison_parts: list[str] = []
    if previous:
        previous_counts = previous.get("counts", {})
        for code in SHEET_ORDER:
            current = counts.get(code)
            prior = previous_counts.get(code)
            if isinstance(current, int) and isinstance(prior, int) and current != prior:
                difference = current - prior
                sign = "+" if difference > 0 else ""
                comparison_parts.append(f"{code} {sign}{difference}")
    comparison = ", ".join(comparison_parts[:8]) if comparison_parts else "비교 가능한 변화 없음"

    history_dir.mkdir(parents=True, exist_ok=True)
    output = history_dir / f"run_{run_date:%Y%m%d}.json"
    payload = {
        "run_date": run_date.isoformat(),
        "created_at": datetime.now(KST).isoformat(),
        "counts": counts,
        "failures": {
            result.config.code: result.error for result in results if not result.success
        },
        "combined": str(combined_path) if combined_path else "",
        "html": str(html_path),
        "pending": str(pending_path) if pending_path else "",
        "comparison": comparison,
    }
    output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return output, comparison


# ---------------------------------------------------------------------------
# 애플리케이션 실행
# ---------------------------------------------------------------------------


class TerminalScheduleApplication:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.run_date = parse_run_date(args.date)
        self.root = Path(args.output_root).expanduser().resolve()
        self.windows_root = PureWindowsPath(args.windows_root)
        self.tools_dir = self.root / "_tools"
        self.log_file = self.tools_dir / "logs" / f"terminal_schedule_{self.run_date:%Y%m%d}.log"
        self.debug_dir = self.tools_dir / "debug" / f"{self.run_date:%Y%m%d}"
        self.browser: BrowserManager | None = None
        self.session = create_requests_session()

    def selected_terminals(self) -> list[TerminalConfig]:
        requested = {
            value.strip().upper()
            for value in (self.args.terminal or "").split(",")
            if value.strip()
        }
        overrides: dict[str, str] = {}
        if self.args.url_overrides:
            override_path = Path(self.args.url_overrides)
            overrides = {
                str(code).upper(): str(url)
                for code, url in json.loads(
                    override_path.read_text(encoding="utf-8")
                ).items()
            }

        selected: list[TerminalConfig] = []
        for config in TERMINALS:
            if requested and config.code not in requested:
                continue
            if config.code in overrides:
                config = TerminalConfig(
                    code=config.code,
                    port=config.port,
                    name=config.name,
                    url=overrides[config.code],
                    folder=config.folder,
                    strategy=config.strategy,
                    expected_columns=config.expected_columns,
                    request_first=config.request_first,
                    note=config.note,
                    menu_texts=config.menu_texts,
                )
            selected.append(config)
        unknown = requested - {config.code for config in TERMINALS}
        if unknown:
            raise CollectorError(f"알 수 없는 터미널 코드: {', '.join(sorted(unknown))}")
        return selected

    async def ensure_browser(self) -> BrowserManager:
        if self.args.no_browser:
            raise CollectorError("--no-browser 옵션으로 브라우저 수집이 비활성화되었습니다.")
        if self.browser is None:
            manager = BrowserManager(
                browser_name=self.args.browser,
                headed=self.args.headed,
                timeout_seconds=self.args.timeout,
                debug_dir=self.debug_dir,
                keep_debug=self.args.keep_debug,
            )
            try:
                await manager.start()
            except Exception:
                await manager.close()
                raise
            self.browser = manager
        return self.browser

    async def collect_one(self, config: TerminalConfig) -> TerminalResult:
        collected_at = datetime.now(KST)
        errors: list[str] = []

        if config.request_first:
            try:
                parsed = collect_with_requests(self.session, config, self.args.timeout)
                parsed = validate_parsed_table(config, parsed)
                return TerminalResult(
                    config=config,
                    success=True,
                    headers=parsed.headers,
                    rows=parsed.rows,
                    note=config.note,
                    source_url=parsed.source or config.url,
                    collected_at=collected_at,
                    method="requests",
                )
            except Exception as exc:
                error = sanitize_error(exc)
                errors.append(f"requests: {error}")
                logging.warning("[%s] requests 실패, 브라우저 재시도: %s", config.code, error)

        try:
            browser = await self.ensure_browser()
            if config.strategy == "legacy_hutchison":
                parsed = await collect_legacy_hutchison(
                    browser, config, max_pages=self.args.max_pages
                )
            elif config.strategy == "nexacro":
                parsed = await collect_nexacro(browser, config)
            elif config.strategy == "div_grid":
                parsed = await collect_div_grid(browser, config)
            elif config.strategy == "bpt":
                parsed = await collect_bpt(
                    browser, config, max_pages=self.args.max_pages
                )
            elif config.strategy == "dongbang_grid":
                parsed = await collect_dongbang_grid(browser, config)
            else:
                parsed = await collect_generic_browser(
                    browser, config, max_pages=self.args.max_pages
                )
            parsed = validate_parsed_table(config, parsed)
            return TerminalResult(
                config=config,
                success=True,
                headers=parsed.headers,
                rows=parsed.rows,
                note=config.note,
                source_url=parsed.source or config.url,
                collected_at=collected_at,
                method="playwright",
            )
        except Exception as exc:
            error = sanitize_error(exc)
            errors.append(f"browser: {error}")
            logging.error("[%s] 수집 실패: %s", config.code, error)
            logging.debug(traceback.format_exc())

        return TerminalResult(
            config=config,
            success=False,
            error=" | ".join(errors) or "알 수 없는 수집 실패",
            note=config.note,
            source_url=config.url,
            collected_at=collected_at,
        )

    async def run(self) -> int:
        self.root.mkdir(parents=True, exist_ok=True)
        configure_logging(self.log_file, self.args.verbose)
        selected = self.selected_terminals()
        logging.info(
            "선석배정현황 수집 시작: 버전=%s, 기준일=%s, 대상=%s",
            APP_VERSION,
            self.run_date,
            ",".join(config.code for config in selected),
        )

        results: list[TerminalResult] = []
        try:
            # 실제 실행 옵션(headed/channel 포함)으로 브라우저를 먼저 띄워 본다.
            # 전역 브라우저 오류가 있으면 정적 사이트 일부만 저장하기 전에 종료한다.
            if not self.args.no_browser:
                logging.info("Playwright 브라우저 실행 사전 점검")
                await self.ensure_browser()
                logging.info("Playwright 브라우저 실행 사전 점검 완료")

            for config in selected:
                result = await self.collect_one(config)
                results.append(result)
                if result.success:
                    try:
                        result.output_file = save_terminal_workbook(
                            result, self.root, self.run_date
                        )
                        logging.info(
                            "[%s] 성공: %d건, 파일=%s, 방식=%s",
                            config.code,
                            result.count,
                            result.output_file,
                            result.method,
                        )
                    except Exception as exc:
                        result.success = False
                        result.error = f"Excel 생성 실패: {sanitize_error(exc)}"
                        logging.error("[%s] %s", config.code, result.error)
        finally:
            if self.browser is not None:
                await self.browser.close()

        successful = [result for result in results if result.success]
        combined_path: Path | None = None
        if successful:
            combined_path = save_combined_workbook(results, self.root, self.run_date)
            logging.info("통합 Excel 생성: %s", combined_path)

        html_path = self.tools_dir / f"report_body_{self.run_date:%Y%m%d}.html"
        create_report_html(results, self.run_date, html_path)
        logging.info("본문 HTML 생성: %s", html_path)

        pending_path: Path | None = None
        if getattr(self.args, "skip_mail", False):
            logging.info(
                "메일 1일 1회 게이트: 오늘치 리포트가 이미 존재하여 발송 지시서(JSON)를 "
                "생성하지 않습니다. (엑셀 데이터만 갱신)"
            )
        elif combined_path is not None:
            pending_path = create_pending_json(
                self.root,
                self.windows_root,
                self.run_date,
                html_path,
                combined_path,
            )
            logging.info("발송 지시서 생성: %s", pending_path)
        else:
            logging.error("성공 터미널이 없어 발송 지시서 JSON을 생성하지 않았습니다.")

        _, comparison = create_history_json(
            results,
            self.run_date,
            self.tools_dir / "history",
            combined_path,
            html_path,
            pending_path,
        )

        success_count = len(successful)
        failed = [result for result in results if not result.success]
        total_rows = sum(result.count for result in successful)
        print()
        print(
            f"수집 결과: {len(results)}개 중 {success_count}개 성공, "
            f"{len(failed)}개 실패, 총 {total_rows:,}건"
        )
        for result in results:
            if result.success:
                print(f"  - {result.config.code}: {result.count:,}건 ({result.method})")
            else:
                print(f"  - {result.config.code}: 실패 - {result.error}")
        print(f"수집시각: {datetime.now(KST):%Y-%m-%d %H:%M} KST")
        print(f"본문 HTML: {html_path}")
        print(f"발송 지시서: {pending_path or '미생성'}")
        print(f"통합 파일: {combined_path or '미생성'}")
        print(
            f"<run-summary>{len(results)}개 터미널 중 {success_count}개 성공, "
            f"총 {total_rows:,}건 수집. 지난 회차 대비 {comparison}.</run-summary>"
        )

        if not successful:
            return 2
        return 1 if failed else 0


# ---------------------------------------------------------------------------
# 자체 점검
# ---------------------------------------------------------------------------


def run_self_test() -> int:
    sample_headers = [
        "No.",
        "선석",
        "선사",
        "모선항차",
        "선사항차",
        "선명",
        "ROUTE",
        "접안방향",
        "반입마감",
        "접안예정",
        "출항예정",
        "작업시작",
        "작업완료",
        "양하",
        "선적",
        "S-H",
        "상태",
    ]
    header_html = "".join(f"<th>{html_lib.escape(value)}</th>" for value in sample_headers)
    data = [
        "1",
        "A1",
        "TST",
        "TEST001",
        "001E/001W",
        "TEST VESSEL",
        "TEST",
        "PORT",
        "2026-08-13 03:00",
        "2026-08-13 13:00",
        "2026-08-14 13:00",
        "",
        "",
        "100",
        "200",
        "",
        "PLANNED",
    ]
    data_html = "".join(f"<td>{html_lib.escape(value)}</td>" for value in data)
    sample = f"<html><body><table><tr>{header_html}</tr><tr>{data_html}</tr></table></body></html>"
    parsed = choose_best_table(
        raw_tables_from_html(sample.encode("utf-8"), "utf-8", "self-test")
    )
    if parsed is None:
        raise AssertionError("표 파서가 샘플 표를 찾지 못했습니다.")
    test_config = next(config for config in TERMINALS if config.code == "PCTC")
    parsed = validate_parsed_table(test_config, parsed)
    if len(parsed.headers) != 17 or len(parsed.rows[0]) != 17:
        raise AssertionError("빈 중간 셀을 포함한 17개 컬럼 정렬 검증 실패")
    if parsed.rows[0][11] != "" or parsed.rows[0][12] != "" or parsed.rows[0][15] != "":
        raise AssertionError("빈 셀 위치가 보존되지 않았습니다.")
    if safe_excel_text("-") != "-" or safe_excel_text("=1+1") != "'=1+1":
        raise AssertionError("엑셀 안전 텍스트 및 단독 하이픈 보존 검증 실패")

    with tempfile.TemporaryDirectory(prefix="terminal_schedule_selftest_") as temp_dir:
        root = Path(temp_dir)
        result = TerminalResult(
            config=test_config,
            success=True,
            headers=parsed.headers,
            rows=parsed.rows,
            source_url="https://example.invalid/test",
            method="self-test",
        )
        result.output_file = save_terminal_workbook(result, root, date(2026, 8, 13))
        combined = save_combined_workbook([result], root, date(2026, 8, 13))
        html_path = create_report_html(
            [result], date(2026, 8, 13), root / "_tools" / "report_body_20260813.html"
        )
        pending = create_pending_json(
            root,
            DEFAULT_WINDOWS_ROOT,
            date(2026, 8, 13),
            html_path,
            combined,
        )
        for path in (result.output_file, combined, html_path, pending):
            if path is None or not path.exists():
                raise AssertionError(f"자체 점검 산출물 생성 실패: {path}")

        verify_excel_sample_format(
            result.output_file,
            test_config.code,
            parsed.headers,
            parsed.rows,
        )
        verify_excel_sample_format(
            combined,
            test_config.code,
            parsed.headers,
            parsed.rows,
        )

    print(
        "SELF-TEST PASSED: 표 정렬, 첨부 샘플 Excel 양식, HTML, "
        "pending JSON 생성 정상"
    )
    return 0


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="국내 컨테이너 터미널 선석배정현황 자동 수집기"
    )
    parser.add_argument(
        "--output-root",
        default=str(default_output_root()),
        help="실제 파일 저장 루트. Windows 기본값: D:\\터미널 스케쥴 정보",
    )
    parser.add_argument(
        "--windows-root",
        default=str(DEFAULT_WINDOWS_ROOT),
        help="pending JSON에 기록할 Windows 루트 경로",
    )
    parser.add_argument(
        "--terminal",
        default="",
        help="수집할 터미널 코드. 쉼표 구분. 예: PNIT,PNC,HPNT",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="기준일자 YYYY-MM-DD. 미지정 시 KST 오늘 날짜",
    )
    parser.add_argument(
        "--browser",
        choices=("chromium", "chrome", "msedge"),
        default="chromium",
        help="Playwright 브라우저 채널",
    )
    parser.add_argument(
        "--headed",
        action="store_true",
        help="브라우저 창을 표시하여 실행",
    )
    parser.add_argument(
        "--no-browser",
        action="store_true",
        help="Playwright 수집 비활성화. requests 대상만 실행",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="사이트별 요청/브라우저 제한시간(초), 기본 60",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=10,
        help="페이지네이션 최대 페이지 수, 기본 10",
    )
    parser.add_argument(
        "--url-overrides",
        default="",
        help='URL 재정의 JSON 파일. 예: {"PNIT":"https://..."}',
    )
    parser.add_argument(
        "--keep-debug",
        action="store_true",
        help="실패 사이트의 HTML/스크린샷을 _tools/debug에 저장",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="상세 로그 출력",
    )
    parser.add_argument(
        "--list-terminals",
        action="store_true",
        help="지원 터미널 목록 출력 후 종료",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="외부 사이트 접속 없이 파서 및 산출물 생성 자체 점검",
    )
    parser.add_argument(
        "--setup",
        "--setup-only",
        dest="setup",
        action="store_true",
        help="필수 패키지와 Playwright Chromium을 자동 설치·점검한 뒤 자체 점검",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {APP_VERSION}",
    )
    parser.add_argument(
        "--no-auto-install",
        action="store_true",
        help="누락 패키지와 Chromium 자동 설치를 금지하고 오류만 표시",
    )
    return parser


async def async_main(args: argparse.Namespace) -> int:
    app = TerminalScheduleApplication(args)
    return await app.run()


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    if args.list_terminals:
        for config in TERMINALS:
            print(
                f"{config.code:5} | {config.port:7} | {config.name:24} | "
                f"{config.strategy} | {config.url}"
            )
        return 0

    try:
        auto_install = not args.no_auto_install

        if args.setup:
            ensure_runtime_environment(
                include_browser=not args.no_browser,
                auto_install=auto_install,
            )
            result = run_self_test()
            if result == 0:
                print("SETUP PASSED: Python 패키지, Chromium, 자체 점검 정상")
            return result

        if args.self_test:
            ensure_runtime_environment(
                include_browser=False,
                auto_install=auto_install,
            )
            return run_self_test()

        # [2026-08-14] 메일 1일 1회 게이트:
        # 오늘치 리포트가 이미 발송(sent/) 또는 대기(pending/)면, 이번 실행(폴백 갱신)은
        # 데이터(엑셀)만 갱신하고 발송 지시서(JSON)는 만들지 않는다(중복 메일 방지).
        _rd = parse_run_date(args.date)
        _tools_dir = Path(args.output_root).expanduser().resolve() / "_tools"
        _today_json = f"report_{_rd:%Y%m%d}.json"
        args.skip_mail = (
            (_tools_dir / "sent" / _today_json).exists()
            or (_tools_dir / "pending" / _today_json).exists()
        )
        if args.skip_mail:
            print(
                f"[메일 게이트] 오늘치 리포트가 이미 존재({_today_json})하여 "
                "이번 실행은 엑셀 데이터만 갱신합니다."
            )
        else:
            # 같은 날짜의 이전 pending JSON이 남아 있으면 새 수집이 끝나기 전에
            # 작업 스케줄러가 불완전 보고서를 발송할 수 있으므로 먼저 대피한다.
            quarantine_existing_pending(args.output_root, args.date)

        # 누락된 Playwright로 12개 터미널이 일괄 실패한 뒤 불완전한 보고서가
        # 생성되지 않도록 실제 사이트 접속 전에 환경 점검을 완료한다.
        ensure_runtime_environment(
            include_browser=not args.no_browser,
            auto_install=auto_install,
        )
        return asyncio.run(async_main(args))
    except KeyboardInterrupt:
        print("사용자 중단", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"치명적 오류: {sanitize_error(exc)}", file=sys.stderr)
        if args.verbose:
            traceback.print_exc()
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
